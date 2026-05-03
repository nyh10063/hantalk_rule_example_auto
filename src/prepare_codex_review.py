#!/usr/bin/env python3
"""Prepare review spreadsheets for Codex/human inspection.

This script does not decide TP/FP and does not generate automatic label
suggestions. It only performs mechanical span checks and adds blank Codex
review columns. Final labels must be written by a human reviewer.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover - handled at runtime for xlsx output.
    openpyxl = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = None  # type: ignore[assignment]

try:
    from .detector.span_utils import DEFAULT_GAP_MARKER, format_span_segments, parse_span_segments
except ImportError:  # pragma: no cover - supports direct script execution.
    from detector.span_utils import DEFAULT_GAP_MARKER, format_span_segments, parse_span_segments

SCHEMA_VERSION = "hantalk_prepare_codex_review_report_v1"
REQUIRED_COLUMNS = {"hit_id", "raw_text", "span_segments"}
LABEL_KEYS = ("tp", "fp", "unclear", "blank", "invalid")
SPAN_PARSE_KEYS = ("parsed", "parse_error", "out_of_bounds", "overlap_or_unsorted")

OUTPUT_COLUMN_ORDER = [
    "hit_id",
    "candidate_index",
    "batch_id",
    "text_id",
    "corpus_domain",
    "source",
    "source_file",
    "source_row_index",
    "source_line_no",
    "origin_e_id",
    "unit_id",
    "unit_type",
    "member_e_ids",
    "canonical_form",
    "group",
    "raw_text",
    "regex_match_text",
    "span_segments",
    "span_extracted_text",
    "span_parse_status",
    "span_parse_note",
    "span_key",
    "span_text",
    "regex_match_span",
    "span_source",
    "component_span_status",
    "component_span_enabled",
    "component_spans",
    "partial_component_spans",
    "partial_span_segments",
    "partial_span_text",
    "matched_component_ids",
    "missing_required_component_ids",
    "applied_bridge_ids",
    "detect_rule_ids",
    "hard_fail_rule_ids",
    "codex_review_label",
    "codex_review_span_status",
    "codex_review_reason",
    "codex_review_note",
    "codex_checked",
    "human_label",
    "span_status",
    "corrected_span_segments",
    "corrected_span_text",
    "memo",
    "reviewer",
    "human_final_check_note",
]

CODEX_REVIEW_COLUMNS = {
    "codex_review_label",
    "codex_review_span_status",
    "codex_review_reason",
    "codex_review_note",
    "codex_checked",
    "human_final_check_note",
}

LABEL_ALIASES = {
    "tp": "tp",
    "true_positive": "tp",
    "true positive": "tp",
    "positive": "tp",
    "pos": "tp",
    "fp": "fp",
    "false_positive": "fp",
    "false positive": "fp",
    "negative": "fp",
    "neg": "fp",
    "unclear": "unclear",
    "unsure": "unclear",
    "?": "unclear",
}


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_json(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2, allow_nan=False)
        f.write("\n")


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _normalize_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return "blank"
    return LABEL_ALIASES.get(text, "invalid")


def _counter_to_counts(counter: Counter[str], keys: tuple[str, ...]) -> dict[str, int]:
    return {key: int(counter.get(key, 0)) for key in keys}


def _validate_headers(path: Path, headers: list[str]) -> list[str]:
    normalized = [_normalize_header(header) for header in headers]
    while normalized and not normalized[-1]:
        normalized.pop()
    if not normalized:
        raise ValueError(f"{path}: missing header row")
    if any(not header for header in normalized):
        raise ValueError(f"{path}: blank header cell inside header row")
    if len(set(normalized)) != len(normalized):
        duplicates = sorted(header for header in set(normalized) if normalized.count(header) > 1)
        raise ValueError(f"{path}: duplicate header after strip: {duplicates}")
    return normalized


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"{path}: missing header row")
        headers = _validate_headers(path, list(reader.fieldnames))
        rows: list[dict[str, str]] = []
        for row in reader:
            rows.append(
                {
                    normalized_name: str(row.get(original_name) or "").strip()
                    for original_name, normalized_name in zip(reader.fieldnames, headers)
                }
            )
        return rows


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read .xlsx files")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError(f"{path}: empty workbook") from None
    headers = _validate_headers(path, ["" if value is None else str(value) for value in header_row])

    rows: list[dict[str, str]] = []
    for values in rows_iter:
        trimmed = list(values[: len(headers)])
        if not any(value is not None and str(value).strip() for value in trimmed):
            continue
        rows.append(
            {
                header: "" if value is None else str(value).strip()
                for header, value in zip(headers, trimmed)
            }
        )
    return rows


def read_review_file(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported review file extension: {path}")


def _validate_required_columns(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        raise ValueError(f"{path}: no data rows")
    columns = set(rows[0])
    missing = sorted(REQUIRED_COLUMNS - columns)
    if missing:
        raise ValueError(f"{path}: missing required columns: {missing}")
    seen_hit_ids: dict[str, int] = {}
    for row_index, row in enumerate(rows, start=2):
        hit_id = str(row.get("hit_id") or "").strip()
        if not hit_id:
            raise ValueError(f"{path}:{row_index} blank hit_id")
        if hit_id in seen_hit_ids:
            raise ValueError(f"{path}:{row_index} duplicate hit_id {hit_id!r}; first row={seen_hit_ids[hit_id]}")
        seen_hit_ids[hit_id] = row_index


def _inspect_span(raw_text: str, value: Any) -> dict[str, Any]:
    try:
        segments = parse_span_segments(value)
    except Exception as exc:
        return {
            "status": "parse_error",
            "note": f"invalid span_segments: {exc}",
            "segments": None,
            "span_extracted_text": "",
            "formatted_span_segments": str(value or "").strip(),
        }

    previous_end: int | None = None
    text_len = len(raw_text)
    for idx, segment in enumerate(segments):
        start, end = int(segment[0]), int(segment[1])
        if start < 0 or end < 0 or start >= end or end > text_len:
            return {
                "status": "out_of_bounds",
                "note": f"segment {idx} is outside [0,{text_len}] or violates start < end: {segment}",
                "segments": segments,
                "span_extracted_text": "",
                "formatted_span_segments": format_span_segments(segments),
            }
        if previous_end is not None and start < previous_end:
            return {
                "status": "overlap_or_unsorted",
                "note": f"segment {idx} overlaps previous segment or is unsorted: {segment}",
                "segments": segments,
                "span_extracted_text": "",
                "formatted_span_segments": format_span_segments(segments),
            }
        previous_end = end

    return {
        "status": "parsed",
        "note": "ok",
        "segments": segments,
        "span_extracted_text": DEFAULT_GAP_MARKER.join(raw_text[start:end] for start, end in segments),
        "formatted_span_segments": format_span_segments(segments),
    }


def _build_output_columns(input_columns: list[str]) -> list[str]:
    ordered: list[str] = []
    for column in OUTPUT_COLUMN_ORDER:
        if column not in ordered:
            ordered.append(column)
    for column in input_columns:
        if column not in ordered:
            ordered.append(column)
    return ordered


def _build_output_row(row: dict[str, str]) -> dict[str, str]:
    raw_text = str(row.get("raw_text") or "")
    span_info = _inspect_span(raw_text, row.get("span_segments"))
    output = dict(row)
    if span_info["status"] == "parsed":
        output["span_segments"] = str(span_info["formatted_span_segments"])
    output["span_extracted_text"] = str(span_info["span_extracted_text"])
    output["span_parse_status"] = str(span_info["status"])
    output["span_parse_note"] = str(span_info["note"])
    for column in CODEX_REVIEW_COLUMNS:
        output.setdefault(column, "")
    return output


def _csv_cell(value: Any) -> str:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
    if value is None:
        return ""
    return str(value)


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_cell(row.get(column, "")) for column in columns})


def write_xlsx(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to write .xlsx files")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "codex_review"
    sheet.append(columns)
    for row in rows:
        sheet.append([_csv_cell(row.get(column, "")) for column in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    parse_error_fill = PatternFill("solid", fgColor="FCE4D6")
    fallback_fill = PatternFill("solid", fgColor="FFF2CC")
    checked_fill = PatternFill("solid", fgColor="E2F0D9")
    header_font = Font(bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_index = {name: idx + 1 for idx, name in enumerate(columns)}
    status_col = column_index.get("span_parse_status")
    span_source_col = column_index.get("span_source")
    checked_col = column_index.get("codex_checked")
    for row_idx in range(2, sheet.max_row + 1):
        fill = None
        if status_col and str(sheet.cell(row_idx, status_col).value or "") != "parsed":
            fill = parse_error_fill
        elif span_source_col and str(sheet.cell(row_idx, span_source_col).value or "") == "regex_match_fallback":
            fill = fallback_fill
        elif checked_col and str(sheet.cell(row_idx, checked_col).value or "").strip().lower() == "yes":
            fill = checked_fill
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill is not None:
                cell.fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    width_overrides = {
        "raw_text": 70,
        "regex_match_text": 18,
        "span_segments": 20,
        "span_extracted_text": 22,
        "span_parse_note": 32,
        "codex_review_reason": 28,
        "codex_review_note": 36,
        "memo": 36,
        "human_final_check_note": 36,
    }
    for idx, column in enumerate(columns, start=1):
        width = width_overrides.get(column, min(max(len(column) + 2, 10), 24))
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    workbook.save(path)


def _count_values(rows: list[dict[str, str]], key: str) -> dict[str, int]:
    counter = Counter(str(row.get(key) or "blank").strip() or "blank" for row in rows)
    return dict(sorted((value, int(count)) for value, count in counter.items()))


def _make_report(
    *,
    item_id: str,
    input_path: Path,
    out_csv: Path,
    out_xlsx: Path,
    report_json: Path,
    rows: list[dict[str, str]],
) -> dict[str, Any]:
    span_parse_counter = Counter(row.get("span_parse_status") or "blank" for row in rows)
    human_label_counter = Counter(_normalize_label(row.get("human_label")) for row in rows)
    return {
        "schema_version": SCHEMA_VERSION,
        "item_id": item_id,
        "created_at": _now_utc(),
        "input": str(input_path),
        "out_csv": str(out_csv),
        "out_xlsx": str(out_xlsx),
        "report_json": str(report_json),
        "n_rows": len(rows),
        "span_parse_counts": _counter_to_counts(span_parse_counter, SPAN_PARSE_KEYS),
        "span_source_counts": _count_values(rows, "span_source"),
        "component_span_status_counts": _count_values(rows, "component_span_status"),
        "existing_human_label_counts": _counter_to_counts(human_label_counter, LABEL_KEYS),
        "required_columns": sorted(REQUIRED_COLUMNS),
        "codex_review_columns": sorted(CODEX_REVIEW_COLUMNS - {"human_final_check_note"}),
        "note": (
            "This script does not generate TP/FP auto suggestions. "
            "Codex and human reviewers must inspect all rows. Final labels must be written by a human reviewer."
        ),
    }


def prepare_codex_review(
    *,
    item_id: str,
    input_path: Path,
    out_csv: Path,
    out_xlsx: Path,
    report_json: Path,
) -> dict[str, Any]:
    item_id = item_id.strip()
    if not item_id:
        raise ValueError("--item-id must not be blank")
    rows = read_review_file(input_path)
    _validate_required_columns(input_path, rows)
    input_columns = list(rows[0].keys())
    output_rows = [_build_output_row(row) for row in rows]
    output_columns = _build_output_columns(input_columns)
    write_csv(out_csv, output_rows, output_columns)
    write_xlsx(out_xlsx, output_rows, output_columns)
    report = _make_report(
        item_id=item_id,
        input_path=input_path,
        out_csv=out_csv,
        out_xlsx=out_xlsx,
        report_json=report_json,
        rows=output_rows,
    )
    _write_json(report_json, report)
    return report


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--item-id", required=True)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--out-csv", required=True, type=Path)
    parser.add_argument("--out-xlsx", required=True, type=Path)
    parser.add_argument("--report-json", required=True, type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)
    try:
        report = prepare_codex_review(
            item_id=args.item_id,
            input_path=args.input,
            out_csv=args.out_csv,
            out_xlsx=args.out_xlsx,
            report_json=args.report_json,
        )
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    print(
        json.dumps(
            {
                "item_id": report["item_id"],
                "n_rows": report["n_rows"],
                "span_parse_counts": report["span_parse_counts"],
                "out_csv": report["out_csv"],
                "out_xlsx": report["out_xlsx"],
                "report_json": report["report_json"],
            },
            ensure_ascii=False,
            indent=2,
            allow_nan=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
