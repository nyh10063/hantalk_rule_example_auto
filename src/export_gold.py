#!/usr/bin/env python3
"""Export human-managed regex gold Excel files to item/unit JSONL.

The source Excel remains the human-managed file. This CLI creates the
machine-friendly JSONL used by test_gold.py.
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from detector.span_utils import make_span_text, parse_span_segments, validate_span_segments
except ImportError:  # pragma: no cover - supports `python -m src.export_gold`.
    from src.detector.span_utils import make_span_text, parse_span_segments, validate_span_segments


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _text(value: Any) -> str:
    if _is_blank(value):
        return ""
    return str(value).strip()


def _split_ids(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing sheet {sheet_name!r} in {path}")
    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    last_non_blank_idx = -1
    for idx, cell in enumerate(header_row):
        if not _is_blank(cell):
            last_non_blank_idx = idx
    if last_non_blank_idx < 0:
        raise ValueError(f"{path}:{sheet_name} has no header row")

    headers: list[str] = []
    seen: set[str] = set()
    for idx, cell in enumerate(header_row[: last_non_blank_idx + 1]):
        if _is_blank(cell):
            raise ValueError(f"{path}:{sheet_name} has blank header at column {idx + 1}")
        header = str(cell).strip()
        if header in seen:
            raise ValueError(f"{path}:{sheet_name} has duplicated header: {header}")
        seen.add(header)
        headers.append(header)

    rows: list[dict[str, Any]] = []
    for row_no, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not values or all(_is_blank(value) for value in values):
            continue
        row = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        row["_row_no"] = row_no
        rows.append(row)
    return headers, rows


def _role_from_pattern_type(pattern_type: str) -> str:
    return "pos_disconti" if pattern_type == "disconti" else "pos_conti"


def _build_record(row: dict[str, Any], *, unit_id: str, example_index: int) -> dict[str, Any]:
    row_no = int(row["_row_no"])
    ps_id = _text(row.get("ps_id") or row.get("unit_id") or row.get("item_id"))
    if ps_id != unit_id:
        raise ValueError(f"gold row {row_no} has ps_id/unit_id/item_id={ps_id!r}, expected {unit_id!r}")

    sentence = _text(row.get("target_sentence") or row.get("sentence"))
    if not sentence:
        raise ValueError(f"gold row {row_no} target_sentence must not be blank")

    span_segments = validate_span_segments(sentence, parse_span_segments(row.get("span_segments")))
    target_text = make_span_text(sentence, span_segments)
    target_spans = [
        {"start": int(start), "end": int(end), "text": sentence[int(start) : int(end)]}
        for start, end in span_segments
    ]
    member_e_ids = _split_ids(row.get("member_e_ids"))
    pattern_type = _text(row.get("pattern_type")) or ("disconti" if len(span_segments) > 1 else "conti")
    example_no = _text(row.get("example_no")) or f"{example_index:03d}"

    return {
        "item_id": unit_id,
        "unit_id": unit_id,
        "member_e_ids": member_e_ids,
        "example_id": f"{unit_id}-GOLD-{example_index:03d}",
        "example_no": example_no,
        "sentence": sentence,
        "target_text": target_text,
        "target_spans": target_spans,
        "source_e_id": ";".join(member_e_ids),
        "source": _text(row.get("source")),
        "split": _text(row.get("split")) or "train",
        "pattern_type": pattern_type,
        "gold_example_role": _text(row.get("gold_example_role")) or _role_from_pattern_type(pattern_type),
        "context_left": _text(row.get("context_left")),
        "context_right": _text(row.get("context_right")),
        "note": _text(row.get("note")) or "converted from regex gold Excel",
    }


def export_gold(
    *,
    input_xlsx: Path,
    out: Path,
    unit_id: str,
    sheet: str = "gold",
    expected_count: int | None = 50,
) -> list[dict[str, Any]]:
    headers, rows = _read_sheet(input_xlsx, sheet)
    required = {"ps_id", "target_sentence", "span_segments"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{input_xlsx}:{sheet} missing required column(s): {', '.join(missing)}")
    records = [_build_record(row, unit_id=unit_id, example_index=idx) for idx, row in enumerate(rows, start=1)]
    if expected_count is not None and len(records) != expected_count:
        raise ValueError(f"Expected {expected_count} gold rows for {unit_id}, found {len(records)}")

    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False, allow_nan=False) + "\n")
    return records


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export HanTalk regex gold Excel to JSONL.")
    parser.add_argument("--input-xlsx", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--unit-id", required=True)
    parser.add_argument("--sheet", default="gold")
    parser.add_argument("--expected-count", type=int, default=50)
    parser.add_argument("--no-expected-count", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    expected_count = None if args.no_expected_count else args.expected_count
    try:
        records = export_gold(
            input_xlsx=args.input_xlsx,
            out=args.out,
            unit_id=args.unit_id,
            sheet=args.sheet,
            expected_count=expected_count,
        )
    except Exception as exc:  # noqa: BLE001 - CLI should print a compact fatal message.
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 2

    print(f"unit_id={args.unit_id}")
    print(f"gold_total={len(records)}")
    print(f"out={args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
