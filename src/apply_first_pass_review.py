#!/usr/bin/env python3
"""Apply Codex first-pass review labels to a prepared review file.

The generated labels are advisory only. This script must not write final
human labels; final TP/FP/span decisions remain in human_label/span_status.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover - handled at runtime for xlsx output.
    openpyxl = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = None  # type: ignore[assignment]

try:
    from .detector.span_utils import parse_span_segments
except ImportError:  # pragma: no cover - supports direct script execution.
    from detector.span_utils import parse_span_segments


SCHEMA_VERSION = "hantalk_codex_first_pass_review_v1"
FIRST_PASS_COLUMNS = [
    "codex_review_label",
    "codex_review_span_status",
    "codex_review_reason",
    "codex_review_note",
]
HUMAN_REVIEW_COLUMNS = [
    "human_label",
    "span_status",
]
BUILTIN_PROFILE_BY_ITEM_ID = {
    "ps_ce002": "ps_ce002_v1",
    "ps_df004": "ps_df004_v1",
}
REASON_LABEL_KO = {
    "auxiliary_go_malda": "보조용언 고 말다 후보",
    "speech_verb_malhada": "말하다/말씀하다 계열 발화 동사",
    "lexical_noun_mal": "명사 말/말이야 표현",
    "tentative_go_malda": "형태상 고 말다 후보",
    "caution_malda3_stop_action": "말다3 중지/그만둠 의미 주의",
    "caution_dep_noun_de_spacing": "의존명사 데 띄어쓰기/STT 혼동 주의",
    "lexicalized_discourse_marker_geureonde": "그런데 계열 접속부사",
    "lexicalized_discourse_marker_geunde": "근데 담화표지/접속부사",
    "place_noun_gunde": "군데 장소/수량 명사",
    "copula_contracted_gunde": "군데/인데 축약 가능",
    "lexical_noun_or_loanword": "가운데/외래어 단어 내부",
    "loanword_pandemic": "팬데믹 단어 내부",
    "loanword_or_title_oneday": "원데이/고유명 표현 내부",
    "lexicalized_ondegande": "온데간데 단어 내부",
    "noisy_or_nonstandard_form": "비표준/오타 가능 후보",
    "productive_connective_ending": "생산적 연결어미 후보",
    "tentative_connective_ending": "형태상 연결어미 후보",
    "unclassified": "미분류 후보",
    "no_first_pass_profile": "1차 검토 profile 없음",
}


def _default_cautions_path(item_id: str) -> Path:
    return Path(__file__).resolve().parents[1] / "configs" / "first_pass_cautions" / f"{item_id}.json"


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_json(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def _reason_label_ko(reason: str) -> str:
    return REASON_LABEL_KO.get(reason, reason or "빈 이유")


def _validate_headers(path: Path, headers: list[str]) -> list[str]:
    normalized = [str(header or "").strip() for header in headers]
    while normalized and not normalized[-1]:
        normalized.pop()
    if not normalized:
        raise ValueError(f"{path}: missing header row")
    if any(not header for header in normalized):
        raise ValueError(f"{path}: blank header cell inside header row")
    duplicates = sorted(header for header in set(normalized) if normalized.count(header) > 1)
    if duplicates:
        raise ValueError(f"{path}: duplicate header after strip: {duplicates}")
    return normalized


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"{path}: missing header row")
        headers = _validate_headers(path, list(reader.fieldnames))
        rows = [
            {
                normalized_name: str(row.get(original_name) or "").strip()
                for original_name, normalized_name in zip(reader.fieldnames, headers)
            }
            for row in reader
        ]
    return headers, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read .xlsx files")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError(f"{path}: empty workbook") from None
    headers = _validate_headers(path, ["" if value is None else str(value) for value in header_row])
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        trimmed = list(values[: len(headers)])
        if not any(value is not None and str(value).strip() for value in trimmed):
            continue
        rows.append({header: "" if value is None else str(value).strip() for header, value in zip(headers, trimmed)})
    return headers, rows


def read_review_file(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported review file extension: {path}")


def _validate_rows(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        raise ValueError(f"{path}: no data rows")
    required = {"hit_id", "raw_text", "regex_match_text", "span_segments"}
    missing = sorted(required - set(rows[0]))
    if missing:
        raise ValueError(f"{path}: missing required columns: {missing}")
    seen_hit_ids: dict[str, int] = {}
    for row_no, row in enumerate(rows, start=2):
        hit_id = str(row.get("hit_id") or "").strip()
        if not hit_id:
            raise ValueError(f"{path}:{row_no} blank hit_id")
        if hit_id in seen_hit_ids:
            raise ValueError(f"{path}:{row_no} duplicate hit_id {hit_id!r}; first row={seen_hit_ids[hit_id]}")
        seen_hit_ids[hit_id] = row_no


def _span_context(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    raw_text = str(row.get("raw_text") or "")
    try:
        segments = parse_span_segments(row.get("span_segments"))
    except Exception:
        return "", "", "", "", ""
    if not segments:
        return "", "", "", "", ""
    start, end = int(segments[0][0]), int(segments[0][1])
    before = raw_text[max(0, start - 6) : start]
    after = raw_text[end : min(len(raw_text), end + 6)]
    window = raw_text[max(0, start - 8) : min(len(raw_text), end + 8)]
    prev = raw_text[start - 1] if start > 0 else ""
    next_ch = raw_text[end] if end < len(raw_text) else ""
    return before, after, window, prev, next_ch


def _regex_match_pair(row: dict[str, str]) -> tuple[int, int] | None:
    raw_text = str(row.get("raw_text") or "")
    match_text = str(row.get("regex_match_text") or "")
    value = str(row.get("regex_match_span") or "").strip()
    if value:
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list) and len(parsed) == 2:
                start, end = int(parsed[0]), int(parsed[1])
                if 0 <= start < end <= len(raw_text):
                    return start, end
        except Exception:
            pass
    if match_text:
        start = raw_text.find(match_text)
        if start >= 0:
            return start, start + len(match_text)
    return None


def _regex_with_context(row: dict[str, str], *, token_window: int = 5) -> str:
    """Return up to N whitespace-token context around regex_match_span."""
    raw_text = str(row.get("raw_text") or "")
    match_pair = _regex_match_pair(row)
    if not raw_text or match_pair is None:
        return ""
    match_start, match_end = match_pair
    tokens = [
        {
            "text": match.group(0),
            "start": match.start(),
            "end": match.end(),
        }
        for match in re.finditer(r"\S+", raw_text)
    ]
    if not tokens:
        return raw_text[match_start:match_end]
    overlapping = [
        idx
        for idx, token in enumerate(tokens)
        if max(match_start, int(token["start"])) < min(match_end, int(token["end"]))
    ]
    if overlapping:
        first_idx = min(overlapping)
        last_idx = max(overlapping)
    else:
        first_idx = next((idx for idx, token in enumerate(tokens) if int(token["end"]) >= match_start), 0)
        last_idx = first_idx
    left_idx = max(0, first_idx - token_window)
    right_idx = min(len(tokens), last_idx + token_window + 1)
    return " ".join(str(token["text"]) for token in tokens[left_idx:right_idx])


def _adjacent_regex_tokens(row: dict[str, str]) -> tuple[str, str]:
    """Return left fragment/token and up to two right fragments/tokens.

    If regex_match_span splits a whitespace token, use the remaining fragment as
    the adjacent token. For example, "보고 말하는 거야?" with regex "고 말"
    yields left="보" and right="하는 거야?".
    """
    raw_text = str(row.get("raw_text") or "")
    match_pair = _regex_match_pair(row)
    if not raw_text or match_pair is None:
        return "", ""
    match_start, match_end = match_pair
    tokens = [
        {
            "text": match.group(0),
            "start": match.start(),
            "end": match.end(),
        }
        for match in re.finditer(r"\S+", raw_text)
    ]
    if not tokens:
        return "", ""
    left = ""
    for token in reversed(tokens):
        token_start = int(token["start"])
        token_end = int(token["end"])
        if token_start < match_start < token_end:
            left = raw_text[token_start:match_start]
            break
        if token_end <= match_start:
            left = str(token["text"])
            break

    right_parts: list[str] = []
    for token in tokens:
        token_start = int(token["start"])
        token_end = int(token["end"])
        if token_start < match_end < token_end:
            suffix = raw_text[match_end:token_end]
            if suffix:
                right_parts.append(suffix)
            continue
        if token_start >= match_end:
            right_parts.append(str(token["text"]))
        if len(right_parts) >= 2:
            break
    right = " ".join(right_parts[:2])
    return left, right


def _load_cautions(item_id: str, cautions_path: Path | None, *, skip_cautions: bool) -> dict[str, Any]:
    if skip_cautions:
        return {
            "status": "skipped_by_flag",
            "source": None,
            "cautions": [],
        }
    path = cautions_path or _default_cautions_path(item_id)
    if not path.exists():
        return {
            "status": "missing",
            "source": str(path),
            "cautions": [],
        }
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"{path}: root must be an object")
    file_item_id = str(data.get("item_id") or "").strip()
    if file_item_id and file_item_id != item_id:
        raise ValueError(f"{path}: item_id={file_item_id!r} does not match --item-id={item_id!r}")
    cautions = data.get("cautions") or []
    if not isinstance(cautions, list):
        raise ValueError(f"{path}: cautions must be a list")
    normalized: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for idx, caution in enumerate(cautions, start=1):
        if not isinstance(caution, dict):
            raise ValueError(f"{path}: caution #{idx} must be an object")
        caution_id = str(caution.get("caution_id") or "").strip()
        if not caution_id:
            raise ValueError(f"{path}: caution #{idx} missing caution_id")
        if caution_id in seen_ids:
            raise ValueError(f"{path}: duplicate caution_id={caution_id!r}")
        seen_ids.add(caution_id)
        match = caution.get("match") or {}
        output = caution.get("output") or {}
        if not isinstance(match, dict):
            raise ValueError(f"{path}: caution_id={caution_id} match must be an object")
        if not isinstance(output, dict):
            raise ValueError(f"{path}: caution_id={caution_id} output must be an object")
        normalized.append({**caution, "caution_id": caution_id})
    normalized.sort(key=lambda item: (int(item.get("priority") or 100), str(item.get("caution_id") or "")))
    return {
        "status": "available",
        "source": str(path),
        "cautions": normalized,
    }


def _regex_search(pattern: Any, text: str, *, field_name: str, caution_id: str) -> bool:
    if pattern is None:
        return True
    try:
        return re.search(str(pattern), text) is not None
    except re.error as exc:
        raise ValueError(f"Invalid caution regex for {caution_id}.{field_name}: {exc}") from exc


def _regex_any(patterns: Any, text: str, *, field_name: str, caution_id: str) -> bool:
    if patterns is None:
        return True
    if not isinstance(patterns, list):
        raise ValueError(f"{caution_id}.{field_name} must be a list")
    if not patterns:
        return True
    return any(_regex_search(pattern, text, field_name=field_name, caution_id=caution_id) for pattern in patterns)


def _matches_caution(row: dict[str, str], caution: dict[str, Any]) -> bool:
    caution_id = str(caution.get("caution_id") or "")
    match = caution.get("match") or {}
    raw_text = str(row.get("raw_text") or "")
    compact_raw_text = re.sub(r"\s+", "", raw_text)
    span_text = str(row.get("span_text") or row.get("regex_match_text") or "")
    regex_match_text = str(row.get("regex_match_text") or "")
    before, after, window, prev, next_ch = _span_context(row)
    haystacks = {
        "raw_text": raw_text,
        "compact_raw_text": compact_raw_text,
        "span_text": span_text,
        "regex_match_text": regex_match_text,
        "window": window,
        "before": before,
        "after": after,
        "prev": prev,
        "next": next_ch,
    }
    for field_name, text in haystacks.items():
        if not _regex_search(match.get(f"{field_name}_regex"), text, field_name=f"{field_name}_regex", caution_id=caution_id):
            return False
        if not _regex_any(
            match.get(f"{field_name}_regex_any"),
            text,
            field_name=f"{field_name}_regex_any",
            caution_id=caution_id,
        ):
            return False
    return True


def _classify_with_cautions(row: dict[str, str], cautions: list[dict[str, Any]]) -> tuple[str, str, str, str, str | None]:
    for caution in cautions:
        if not _matches_caution(row, caution):
            continue
        output = caution.get("output") or {}
        caution_id = str(caution.get("caution_id") or "")
        label = str(output.get("codex_review_label") or "unclear").strip()
        span_status = str(output.get("codex_review_span_status") or "ok").strip()
        reason = str(output.get("codex_review_reason") or f"caution_{caution_id}").strip()
        note = str(output.get("codex_review_note") or caution.get("description") or "").strip()
        label_ko = str(caution.get("label_ko") or caution_id).strip()
        if note:
            note = f"주의 유형: {label_ko}. {note}"
        else:
            note = f"주의 유형: {label_ko}"
        return label, span_status, reason, note, caution_id
    return "", "", "", "", None


def _classify_ps_ce002(row: dict[str, str]) -> tuple[str, str, str, str]:
    raw_text = str(row.get("raw_text") or "")
    span_text = str(row.get("span_text") or row.get("regex_match_text") or "")
    before, after, window, _prev, next_ch = _span_context(row)
    compact = raw_text.replace(" ", "")

    if span_text == "런데" and "그런데" in window:
        return "fp", "not_applicable", "lexicalized_discourse_marker_geureonde", "그런데/그런데도 계열 접속부사로 보임"
    if span_text == "근데":
        return "fp", "not_applicable", "lexicalized_discourse_marker_geunde", "근데 담화표지/접속부사로 보임"
    if span_text == "군데":
        if re.search(r"(몇|한|두|세|네|\d+)\s*군데", raw_text) or next_ch in {"가", "를", "에", "도", " "}:
            return "fp", "not_applicable", "place_noun_gunde", "군데 장소/수량 명사로 보임"
        return "tp", "ok", "copula_contracted_gunde", "친군데 등 -인데 축약 가능성이 큼"
    if span_text == "운데" and ("가운데" in compact or "파운데이션" in compact):
        return "fp", "not_applicable", "lexical_noun_or_loanword", "가운데/파운데이션 등 어미가 아닌 단어 내부로 보임"
    if span_text == "팬데" and "팬데믹" in raw_text:
        return "fp", "not_applicable", "loanword_pandemic", "팬데믹 단어 내부"
    if span_text == "원데" and ("원데이" in raw_text or " 원" in before + span_text + after):
        return "fp", "not_applicable", "loanword_or_title_oneday", "원데이/고유명 표현 내부 가능성이 큼"
    if span_text == "온데" and "온데간데" in raw_text:
        return "fp", "not_applicable", "lexicalized_ondegande", "온데간데 단어 내부"
    if span_text in {"천데", "끈데"}:
        return "unclear", "ok", "noisy_or_nonstandard_form", "학습자/구어 비표준 표기 또는 담화표지 오타 가능성 있어 확인 필요"

    productive_spans = {
        "는데",
        "은데",
        "인데",
        "한데",
        "건데",
        "텐데",
        "닌데",
        "큰데",
        "신데",
        "던데",
        "언데",
        "든데",
        "긴데",
        "땐데",
        "뭔데",
        "젠데",
        "딘데",
        "븐데",
    }
    if span_text in productive_spans:
        return "tp", "ok", "productive_connective_ending", "ㄴ/은/는데 계열 연결어미 또는 이다/용언 결합으로 보임"
    if span_text.endswith("데"):
        return "tp", "ok", "tentative_connective_ending", "형태상 ㄴ/은/는데 계열 후보로 보이나 최종 확인 필요"
    return "unclear", "ok", "unclassified", "자동 규칙으로 분류하지 못함"


def _classify_ps_df004(row: dict[str, str]) -> tuple[str, str, str, str]:
    raw_text = str(row.get("raw_text") or "")
    regex_match_text = str(row.get("regex_match_text") or "")
    span_text = str(row.get("span_text") or regex_match_text)
    before, after, window, _prev, next_ch = _span_context(row)
    compact = re.sub(r"\s+", "", raw_text)

    # "고 말했다/말씀/말하는" are frequent systematic FPs for the broad
    # bootstrap pattern 고(?:야)?\s*말. Keep this as advisory first-pass
    # labeling; final human_label remains the source of truth.
    speech_after = re.match(r"(했|했다|했고|했는|하며|해|하면|하기|하긴|하는|전했|씀|을\b)", after)
    if speech_after or re.search(r"고\s*말(?:씀|하|했|해|전했|을\s+하)", window):
        return "fp", "not_applicable", "speech_verb_malhada", "인용/발화 동사 '말하다/말씀하다' 계열로 보임"

    if re.match(r"(이야|이\s|이[가-힣])", after):
        return "fp", "not_applicable", "lexical_noun_mal", "명사 '말' 또는 '말이야' 담화 표현으로 보임"

    if re.search(r"고(?:야)?말(?:았|겠|거|다|고야)", compact) or re.match(r"(았|겠|거|다)", after):
        return "tp", "ok", "auxiliary_go_malda", "보조용언 '고 말다'의 완료/의지 표현으로 보임"

    if span_text in {"고 말", "고 ... 말"} or regex_match_text.startswith("고"):
        return "unclear", "ok", "tentative_go_malda", "형태상 고 말다 후보이나 말하다/명사 말 가능성 확인 필요"

    return "unclear", "ok", "unclassified", "자동 규칙으로 분류하지 못함"


def _classify_no_profile(row: dict[str, str]) -> tuple[str, str, str, str]:
    return "", "", "no_first_pass_profile", "이 unit에 대한 Codex 1차 검토 profile이 아직 없음"


def _classify_row(row: dict[str, str], profile_id: str) -> tuple[str, str, str, str]:
    if profile_id == "ps_ce002_v1":
        return _classify_ps_ce002(row)
    if profile_id == "ps_df004_v1":
        return _classify_ps_df004(row)
    if profile_id == "none":
        return _classify_no_profile(row)
    raise ValueError(f"Unsupported first-pass profile_id={profile_id!r}")


def _build_output_columns(input_columns: list[str]) -> list[str]:
    review_columns = [
        "human_label",
        "span_status",
        "codex_review_label",
        "regex_match_text",
        "a_token_left",
        "two_tokens_right",
        "codex_review_span_status",
        "codex_review_reason",
        "codex_review_note",
    ]
    generated_columns = set(review_columns) | {"regex_with_context", "a_token_right"}
    columns = [column for column in input_columns if column not in generated_columns]
    if "raw_text" in columns:
        insert_at = columns.index("raw_text") + 1
        columns = columns[:insert_at] + ["regex_with_context"] + columns[insert_at:]
        return columns[: insert_at + 1] + review_columns + columns[insert_at + 1 :]
    return ["regex_with_context"] + review_columns + columns


def _csv_cell(value: Any) -> str:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
    if value is None:
        return ""
    return str(value)


def _write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_cell(row.get(column, "")) for column in columns})


def _write_xlsx(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to write .xlsx files")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "codex_first_pass"
    sheet.append(columns)
    for row in rows:
        sheet.append([_csv_cell(row.get(column, "")) for column in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    tp_fill = PatternFill("solid", fgColor="E2F0D9")
    fp_fill = PatternFill("solid", fgColor="FCE4D6")
    unclear_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_index = {name: idx + 1 for idx, name in enumerate(columns)}
    label_col = column_index.get("codex_review_label")
    for row_idx in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row_idx, label_col).value or "") if label_col else ""
        fill = tp_fill if label == "tp" else fp_fill if label == "fp" else unclear_fill if label == "unclear" else None
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill is not None:
                cell.fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    width_overrides = {
        "raw_text": 70,
        "regex_with_context": 70,
        "regex_match_text": 18,
        "a_token_left": 18,
        "two_tokens_right": 24,
        "codex_review_label": 18,
        "codex_review_span_status": 22,
        "codex_review_reason": 32,
        "codex_review_note": 52,
        "span_segments": 20,
        "memo": 36,
    }
    for idx, column in enumerate(columns, start=1):
        width = width_overrides.get(column, min(max(len(column) + 2, 10), 24))
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    workbook.save(path)


def apply_first_pass_review(
    *,
    item_id: str,
    input_path: Path,
    out_csv: Path,
    out_xlsx: Path,
    report_json: Path,
    profile_id: str | None = None,
    cautions_path: Path | None = None,
    skip_cautions: bool = False,
) -> dict[str, Any]:
    item_id = item_id.strip()
    if not item_id:
        raise ValueError("--item-id must not be blank")
    profile = profile_id or BUILTIN_PROFILE_BY_ITEM_ID.get(item_id, "none")
    profile_status = "missing" if profile == "none" else "available"
    if profile_id:
        profile_source = "cli"
    elif item_id in BUILTIN_PROFILE_BY_ITEM_ID:
        profile_source = "builtin"
    else:
        profile_source = "none"
    caution_config = _load_cautions(item_id, cautions_path, skip_cautions=skip_cautions)
    cautions = caution_config["cautions"]
    advisory_labels_applied = profile_status == "available" or bool(cautions)
    input_columns, rows = read_review_file(input_path)
    _validate_rows(input_path, rows)
    output_columns = _build_output_columns(input_columns)
    output_rows: list[dict[str, str]] = []
    label_counts: Counter[str] = Counter()
    span_status_counts: Counter[str] = Counter()
    reason_counts: Counter[str] = Counter()
    reason_ko_counts: Counter[str] = Counter()
    caution_counts: Counter[str] = Counter()
    examples_by_reason_code: dict[str, list[dict[str, str]]] = defaultdict(list)
    examples_by_reason_ko: dict[str, list[dict[str, str]]] = defaultdict(list)
    examples_by_caution: dict[str, list[dict[str, str]]] = defaultdict(list)

    for row in rows:
        out_row = dict(row)
        label, span_status, reason, note, caution_id = _classify_with_cautions(out_row, cautions)
        if caution_id is None:
            label, span_status, reason, note = _classify_row(out_row, profile)
        out_row["regex_with_context"] = _regex_with_context(out_row)
        a_token_left, two_tokens_right = _adjacent_regex_tokens(out_row)
        out_row["a_token_left"] = a_token_left
        out_row["two_tokens_right"] = two_tokens_right
        out_row["codex_review_label"] = label
        out_row["codex_review_span_status"] = span_status
        out_row["codex_review_reason"] = reason
        out_row["codex_review_note"] = (
            f"{note} / Codex 1차 검토이며 최종 라벨은 human_label 기준" if note else ""
        )
        out_row["codex_checked"] = "yes" if label or reason != "no_first_pass_profile" else ""
        output_rows.append(out_row)

        label_counts[label or "blank"] += 1
        span_status_counts[span_status or "blank"] += 1
        reason_counts[reason] += 1
        reason_ko = _reason_label_ko(reason)
        reason_ko_counts[reason_ko] += 1
        if caution_id is not None:
            caution_counts[caution_id] += 1
        example = {
            "hit_id": out_row.get("hit_id", ""),
            "span_text": out_row.get("span_text", ""),
            "raw_text": out_row.get("raw_text", ""),
            "codex_review_label": label,
            "codex_review_reason_code": reason,
            "codex_review_reason_ko": reason_ko,
        }
        if caution_id is not None:
            example["caution_id"] = caution_id
        if len(examples_by_reason_code[reason]) < 5:
            examples_by_reason_code[reason].append(example)
        if len(examples_by_reason_ko[reason_ko]) < 5:
            examples_by_reason_ko[reason_ko].append(example)
        if caution_id is not None and len(examples_by_caution[caution_id]) < 10:
            examples_by_caution[caution_id].append(example)

    _write_csv(out_csv, output_rows, output_columns)
    _write_xlsx(out_xlsx, output_rows, output_columns)
    report = {
        "schema_version": SCHEMA_VERSION,
        "item_id": item_id,
        "profile_id": profile,
        "profile_status": profile_status,
        "profile_source": profile_source,
        "advisory_labels_applied": advisory_labels_applied,
        "created_at": _now_utc(),
        "input": str(input_path),
        "out_csv": str(out_csv),
        "out_xlsx": str(out_xlsx),
        "report_json": str(report_json),
        "n_rows": len(output_rows),
        "codex_review_label_counts": dict(sorted(label_counts.items())),
        "codex_review_span_status_counts": dict(sorted(span_status_counts.items())),
        "codex_review_reason_counts": dict(reason_counts.most_common()),
        "codex_review_reason_ko_counts": dict(reason_ko_counts.most_common()),
        "examples_by_reason": dict(examples_by_reason_ko),
        "examples_by_reason_code": dict(examples_by_reason_code),
        "cautions": {
            "status": caution_config["status"],
            "source": caution_config["source"],
            "n_cautions": len(cautions),
            "hit_counts": dict(caution_counts.most_common()),
            "examples_by_caution": dict(examples_by_caution),
            "policy": {
                "applied_before_profile": True,
                "advisory_only": True,
                "human_label_is_final": True,
            },
        },
        "reason_label_policy": {
            "codex_review_reason": "stable_machine_code",
            "examples_by_reason": "korean_human_readable_key",
            "examples_by_reason_code": "stable_machine_code_key",
        },
        "column_policy": {
            "regex_with_context_after": "raw_text",
            "regex_match_text_after": "codex_review_label",
            "adjacent_token_columns_after": "regex_match_text",
            "adjacent_token_columns": ["a_token_left", "two_tokens_right"],
            "human_review_columns_after": "regex_with_context",
            "human_review_columns": HUMAN_REVIEW_COLUMNS,
            "first_pass_columns_after": "span_status",
            "first_pass_columns": FIRST_PASS_COLUMNS,
        },
        "policy": {
            "human_label_is_final": True,
            "codex_first_pass_is_reference_only": True,
            "human_label_columns_modified": False,
            "missing_profile_is_non_blocking": True,
        },
    }
    _write_json(report_json, report)
    return report


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--item-id", required=True)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--out-csv", required=True, type=Path)
    parser.add_argument("--out-xlsx", required=True, type=Path)
    parser.add_argument("--report-json", required=True, type=Path)
    parser.add_argument("--profile-id", help="First-pass profile id. Defaults to ps_ce002_v1 for ps_ce002, otherwise none.")
    parser.add_argument(
        "--cautions",
        type=Path,
        help="Optional first-pass caution repository JSON. Defaults to configs/first_pass_cautions/{item_id}.json if present.",
    )
    parser.add_argument(
        "--skip-cautions",
        action="store_true",
        help="Disable first-pass caution repository lookup for this run.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    try:
        report = apply_first_pass_review(
            item_id=args.item_id,
            input_path=args.input,
            out_csv=args.out_csv,
            out_xlsx=args.out_xlsx,
            report_json=args.report_json,
            profile_id=args.profile_id,
            cautions_path=args.cautions,
            skip_cautions=args.skip_cautions,
        )
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    print(
        json.dumps(
            {
                "item_id": report["item_id"],
                "profile_id": report["profile_id"],
                "profile_status": report["profile_status"],
                "n_rows": report["n_rows"],
                "codex_review_label_counts": report["codex_review_label_counts"],
                "cautions": report["cautions"],
                "out_csv": report["out_csv"],
                "out_xlsx": report["out_xlsx"],
                "report_json": report["report_json"],
            },
            ensure_ascii=False,
            indent=2,
            allow_nan=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
