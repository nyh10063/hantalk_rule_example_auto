#!/usr/bin/env python3
"""Export a runtime detector bundle from the human-managed dict.xlsx."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from .bridges import BRIDGE_REGISTRY, bridge_metadata_by_id
except ImportError:  # pragma: no cover - supports direct script execution.
    from src.detector.bridges import BRIDGE_REGISTRY, bridge_metadata_by_id

SCHEMA_VERSION = "hantalk_detector_bundle_v1"

REQUIRED_SHEETS = {"items", "rule_components", "detect_rules"}
REQUIRED_COLUMNS = {
    "items": {
        "e_id",
        "canonical_form",
        "group",
        "gloss",
    },
    "rule_components": {
        "comp_surf",
        "comp_id",
        "is_required",
        "anchor_rank",
        "comp_order",
        "order_policy",
        "min_gap_to_next",
        "max_gap_to_next",
    },
    "detect_rules": {
        "ruleset_id",
        "rule_id",
        "stage",
        "target",
        "pattern",
        "priority",
        "hard_fail",
    },
}

VALID_GROUPS = {"a", "b", "c"}
VALID_STAGES = {"detect", "verify"}
COMPONENT_SCOPED_TARGETS = {
    "component_right_context",
    "component_left_context",
    "component_text",
    "left_plus_component_text",
}
VALID_TARGETS_BY_STAGE = {
    "detect": {"raw_sentence"},
    "verify": {"raw_sentence", "char_window", *COMPONENT_SCOPED_TARGETS},
}
VALID_ORDER_POLICIES = {"fx", "fl"}
TASK_UNIT_ID_COLUMNS = ("unit_id", "ps_id", "e_id")


class BundleExportError(ValueError):
    """Fatal bundle export error."""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _text(value: Any, *, lower: bool = False, none_if_blank: bool = True) -> str | None:
    if _is_blank(value):
        return None if none_if_blank else ""
    out = str(value).strip()
    return out.lower() if lower else out


def _required_text(record: dict[str, Any], key: str, *, sheet: str, row_no: int, lower: bool = False) -> str:
    value = _text(record.get(key), lower=lower)
    if value is None:
        raise BundleExportError(f"{sheet}:{row_no} missing required value: {key}")
    return value


def _int_or_none(value: Any, *, sheet: str, row_no: int, key: str) -> int | None:
    if _is_blank(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise BundleExportError(f"{sheet}:{row_no} {key} must be int-like: {value!r}") from exc


def _bool_value(value: Any, *, sheet: str, row_no: int, key: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return bool(int(value))
    text = _text(value, lower=True)
    if text is None:
        return False
    if text in {"true", "t", "yes", "y", "1"}:
        return True
    if text in {"false", "f", "no", "n", "0"}:
        return False
    raise BundleExportError(f"{sheet}:{row_no} {key} must be boolean-like: {value!r}")


def _split_ids(value: Any) -> list[str]:
    text = _text(value)
    if text is None:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _task_unit_id(record: dict[str, Any]) -> str | None:
    """Return the explicit runtime task unit id for rows that may be item- or ps-based."""
    for key in TASK_UNIT_ID_COLUMNS:
        value = _text(record.get(key))
        if value is not None:
            return value
    return None


def _required_task_unit_id(record: dict[str, Any], *, sheet: str, row_no: int) -> str:
    unit_id = _task_unit_id(record)
    if unit_id is None:
        keys = "/".join(TASK_UNIT_ID_COLUMNS)
        raise BundleExportError(f"{sheet}:{row_no} missing required task unit id ({keys})")
    return unit_id


def _read_sheet(workbook: Any, sheet_name: str) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    last_non_blank_idx = -1
    for idx, cell in enumerate(header_row):
        if not _is_blank(cell):
            last_non_blank_idx = idx
    if last_non_blank_idx < 0:
        raise BundleExportError(f"{sheet_name} has no header row")

    headers: list[str] = []
    seen_headers: set[str] = set()
    for idx, cell in enumerate(header_row[: last_non_blank_idx + 1]):
        if _is_blank(cell):
            raise BundleExportError(f"{sheet_name} has blank header in the middle at column {idx + 1}")
        header = str(cell).strip()
        if header in seen_headers:
            raise BundleExportError(f"{sheet_name} has duplicated header: {header}")
        seen_headers.add(header)
        headers.append(header)

    records: list[tuple[int, dict[str, Any]]] = []
    for row_no, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not values or all(_is_blank(value) for value in values):
            continue
        record = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        if all(_is_blank(record.get(header)) for header in headers):
            continue
        records.append((row_no, record))
    return headers, records


def _check_required_structure(workbook: Any) -> None:
    missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing_sheets:
        raise BundleExportError(f"Missing required sheet(s): {', '.join(missing_sheets)}")

    for sheet_name, required_columns in REQUIRED_COLUMNS.items():
        headers, _ = _read_sheet(workbook, sheet_name)
        missing_columns = sorted(required_columns - set(headers))
        if missing_columns:
            raise BundleExportError(f"{sheet_name} missing required column(s): {', '.join(missing_columns)}")
        if sheet_name in {"rule_components", "detect_rules"} and not (set(headers) & set(TASK_UNIT_ID_COLUMNS)):
            keys = ", ".join(TASK_UNIT_ID_COLUMNS)
            raise BundleExportError(f"{sheet_name} must contain one task unit id column: {keys}")

    item_headers, _ = _read_sheet(workbook, "items")
    if not ({"ps_id", "polyset_id"} & set(item_headers)):
        # Single-item dictionaries do not need ps_id/polyset_id, so this is not fatal.
        return


def _looks_like_python_regex_literal(pattern: str) -> bool:
    stripped = pattern.strip()
    if re.match(r"^[rRuUfFbB]*(['\"]).*\1$", stripped):
        return True
    return stripped.startswith(("r\"", "r'", "R\"", "R'"))


def _derive_polyset_form(items: list[dict[str, Any]]) -> str:
    forms = [str(item.get("canonical_form") or item["e_id"]) for item in items]
    stripped = [re.sub(r"\d+$", "", form).strip() for form in forms]
    return stripped[0] if stripped and all(form == stripped[0] for form in stripped) else "/".join(forms)


def _join_encoder_gloss(gloss_intro: str | None, member_items: list[dict[str, Any]]) -> str | None:
    member_glosses = [str(item.get("gloss") or "").strip().rstrip(".") for item in member_items if item.get("gloss")]
    member_glosses = [gloss for gloss in member_glosses if gloss]
    intro = (gloss_intro or "").strip()
    if not member_glosses:
        return intro or None
    joined = "; ".join(member_glosses)
    if intro:
        return f"{intro} {joined}"
    return joined


def _polyset_component_ids(row: dict[str, Any]) -> list[str]:
    return (
        _split_ids(row.get("ps_comp_id"))
        or _split_ids(row.get("component_ids"))
        or _split_ids(row.get("e_comp_id"))
    )


def _validate_ruleset_unit(
    *,
    owner_kind: str,
    owner_id: str,
    ruleset_id: str,
    expected_unit_id: str,
    expected_stage: str,
    rules_by_ruleset_id: dict[str, list[dict[str, Any]]],
) -> None:
    rules = [rule for rule in rules_by_ruleset_id.get(ruleset_id, []) if rule.get("stage") == expected_stage]
    if not rules:
        raise BundleExportError(
            f"{owner_kind} {owner_id} {expected_stage}_ruleset_id has no {expected_stage} rules: {ruleset_id}"
        )
    wrong_units = sorted({str(rule.get("unit_id") or rule.get("e_id")) for rule in rules if str(rule.get("unit_id") or rule.get("e_id")) != expected_unit_id})
    if wrong_units:
        raise BundleExportError(
            f"{owner_kind} {owner_id} {expected_stage}_ruleset_id={ruleset_id} contains rules for other unit_id values: "
            f"{', '.join(wrong_units)}; expected {expected_unit_id}"
        )


def build_bundle(dict_xlsx: Path) -> dict[str, Any]:
    workbook = load_workbook(dict_xlsx, read_only=True, data_only=True)
    _check_required_structure(workbook)
    warnings: list[str] = []

    _, item_rows = _read_sheet(workbook, "items")
    _, component_rows = _read_sheet(workbook, "rule_components")
    _, rule_rows = _read_sheet(workbook, "detect_rules")
    polyset_rows: list[tuple[int, dict[str, Any]]] = []
    if "polysets" in workbook.sheetnames:
        polyset_headers, polyset_rows = _read_sheet(workbook, "polysets")
        required_polyset_columns = {"ps_id", "primary_e_id", "member_e_ids", "ps_canonical_form", "gloss_intro", "note"}
        missing_polyset_columns = sorted(required_polyset_columns - set(polyset_headers))
        if missing_polyset_columns:
            raise BundleExportError(
                f"polysets missing required column(s): {', '.join(missing_polyset_columns)}"
            )

    items_by_e_id: dict[str, dict[str, Any]] = {}
    for row_no, row in item_rows:
        e_id = _required_text(row, "e_id", sheet="items", row_no=row_no)
        if e_id in items_by_e_id:
            raise BundleExportError(f"items.e_id duplicated: {e_id}")
        group = _required_text(row, "group", sheet="items", row_no=row_no, lower=True)
        if group not in VALID_GROUPS:
            raise BundleExportError(f"items:{row_no} group must be one of a/b/c: {group!r}")
        ps_id = _text(row.get("ps_id")) or _text(row.get("polyset_id"))
        if group == "c" and ps_id is None:
            warnings.append(f"items:{row_no} group=c but ps_id/polyset_id is empty for e_id={e_id}")
        items_by_e_id[e_id] = {
            "e_id": e_id,
            "canonical_form": _required_text(row, "canonical_form", sheet="items", row_no=row_no),
            "group": group,
            "ps_id": ps_id,
            "polyset_id": ps_id,
            "level": _text(row.get("난이도")),
            "topic": _text(row.get("주제")),
            "disconti_allowed": _bool_value(row.get("disconti_allowed"), sheet="items", row_no=row_no, key="disconti_allowed"),
            "e_comp_ids": _split_ids(row.get("e_comp_id")),
            "detect_ruleset_id": _text(row.get("detect_ruleset_id")),
            "verify_ruleset_id": _text(row.get("verify_ruleset_id")),
            "gloss": _text(row.get("gloss")),
        }

    explicit_polysets_by_id: dict[str, dict[str, Any]] = {}
    for row_no, row in polyset_rows:
        ps_id = _required_text(row, "ps_id", sheet="polysets", row_no=row_no)
        if ps_id in explicit_polysets_by_id:
            raise BundleExportError(f"polysets.ps_id duplicated: {ps_id}")
        member_e_ids = _split_ids(row.get("member_e_ids"))
        if not member_e_ids:
            raise BundleExportError(f"polysets:{row_no} member_e_ids must not be blank for ps_id={ps_id}")
        missing_members = [e_id for e_id in member_e_ids if e_id not in items_by_e_id]
        if missing_members:
            raise BundleExportError(
                f"polysets:{row_no} member_e_ids not found in items for ps_id={ps_id}: {', '.join(missing_members)}"
            )
        primary_e_id = _required_text(row, "primary_e_id", sheet="polysets", row_no=row_no)
        if primary_e_id not in member_e_ids:
            raise BundleExportError(
                f"polysets:{row_no} primary_e_id={primary_e_id} is not in member_e_ids for ps_id={ps_id}"
            )
        for member_e_id in member_e_ids:
            item_ps_id = items_by_e_id[member_e_id].get("ps_id")
            if item_ps_id and item_ps_id != ps_id:
                raise BundleExportError(
                    f"polysets:{row_no} member e_id={member_e_id} has items.ps_id={item_ps_id}, expected {ps_id}"
                )
            if items_by_e_id[member_e_id].get("group") != "c":
                raise BundleExportError(
                    f"polysets:{row_no} member e_id={member_e_id} must have group=c for ps_id={ps_id}"
                )
        member_items = [items_by_e_id[e_id] for e_id in member_e_ids]
        gloss_intro = _text(row.get("gloss_intro"))
        explicit_polysets_by_id[ps_id] = {
            "ps_id": ps_id,
            "polyset_id": ps_id,
            "primary_e_id": primary_e_id,
            "member_e_ids": member_e_ids,
            "canonical_form": _required_text(row, "ps_canonical_form", sheet="polysets", row_no=row_no),
            "disconti_allowed": _bool_value(row.get("disconti_allowed"), sheet="polysets", row_no=row_no, key="disconti_allowed"),
            "ps_comp_ids": _polyset_component_ids(row),
            "gloss_intro": gloss_intro,
            "encoder_gloss": _join_encoder_gloss(gloss_intro, member_items),
            "note": _text(row.get("note")),
            "detect_ruleset_id": _text(row.get("detect_ruleset_id")),
            "verify_ruleset_id": _text(row.get("verify_ruleset_id")),
        }

    components_by_e_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    referenced_bridge_ids: set[str] = set()
    for row_no, row in component_rows:
        unit_id = _required_task_unit_id(row, sheet="rule_components", row_no=row_no)
        if unit_id not in items_by_e_id and unit_id not in explicit_polysets_by_id:
            raise BundleExportError(f"rule_components:{row_no} task unit id not found in items/polysets: {unit_id}")
        bridge_id = _text(row.get("bridge_id"))
        if bridge_id is not None:
            if bridge_id not in BRIDGE_REGISTRY:
                known = ", ".join(sorted(BRIDGE_REGISTRY))
                raise BundleExportError(
                    f"rule_components:{row_no} unknown bridge_id={bridge_id!r}; known bridge_id values: {known}"
                )
            referenced_bridge_ids.add(bridge_id)
        order_policy = _text(row.get("order_policy"), lower=True) or "fx"
        if order_policy not in VALID_ORDER_POLICIES:
            allowed = ", ".join(sorted(VALID_ORDER_POLICIES))
            raise BundleExportError(
                f"rule_components:{row_no} order_policy={order_policy!r} invalid; allowed: {allowed}"
            )
        component = {
            "e_id": unit_id,
            "unit_id": unit_id,
            "source_e_id": _text(row.get("e_id")),
            "ps_id": _text(row.get("ps_id")),
            "comp_surf": _required_text(row, "comp_surf", sheet="rule_components", row_no=row_no),
            "comp_id": _required_text(row, "comp_id", sheet="rule_components", row_no=row_no),
            "is_required": _bool_value(row.get("is_required"), sheet="rule_components", row_no=row_no, key="is_required"),
            "anchor_rank": _int_or_none(row.get("anchor_rank"), sheet="rule_components", row_no=row_no, key="anchor_rank"),
            "comp_order": _int_or_none(row.get("comp_order"), sheet="rule_components", row_no=row_no, key="comp_order"),
            "order_policy": order_policy,
            "min_gap_to_next": _int_or_none(row.get("min_gap_to_next"), sheet="rule_components", row_no=row_no, key="min_gap_to_next"),
            "max_gap_to_next": _int_or_none(row.get("max_gap_to_next"), sheet="rule_components", row_no=row_no, key="max_gap_to_next"),
            "bridge_id": bridge_id,
        }
        components_by_e_id[unit_id].append(component)

    for components in components_by_e_id.values():
        components.sort(key=lambda item: (item.get("comp_order") is None, item.get("comp_order") or 0, item["comp_id"]))

    component_ids_by_e_id = {
        e_id: {str(component["comp_id"]) for component in components}
        for e_id, components in components_by_e_id.items()
    }

    rules_by_ruleset_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen_rule_ids: set[str] = set()
    for row_no, row in rule_rows:
        unit_id = _required_task_unit_id(row, sheet="detect_rules", row_no=row_no)
        if unit_id not in items_by_e_id and unit_id not in explicit_polysets_by_id:
            raise BundleExportError(f"detect_rules:{row_no} task unit id not found in items/polysets: {unit_id}")
        ruleset_id = _required_text(row, "ruleset_id", sheet="detect_rules", row_no=row_no)
        rule_id = _required_text(row, "rule_id", sheet="detect_rules", row_no=row_no)
        if rule_id in seen_rule_ids:
            raise BundleExportError(f"detect_rules.rule_id duplicated: {rule_id}")
        seen_rule_ids.add(rule_id)

        stage = _required_text(row, "stage", sheet="detect_rules", row_no=row_no, lower=True)
        if stage not in VALID_STAGES:
            raise BundleExportError(f"detect_rules:{row_no} stage must be detect/verify: {stage!r}")
        target = _required_text(row, "target", sheet="detect_rules", row_no=row_no, lower=True)
        if target not in VALID_TARGETS_BY_STAGE[stage]:
            allowed = ", ".join(sorted(VALID_TARGETS_BY_STAGE[stage]))
            raise BundleExportError(f"detect_rules:{row_no} target={target!r} invalid for stage={stage}; allowed: {allowed}")
        component_id = _text(row.get("component_id"))
        if target in COMPONENT_SCOPED_TARGETS:
            if component_id is None:
                raise BundleExportError(
                    f"detect_rules:{row_no} target={target} requires component_id for rule_id={rule_id}"
                )
            if component_id not in component_ids_by_e_id.get(unit_id, set()):
                known = ", ".join(sorted(component_ids_by_e_id.get(unit_id, set())))
                raise BundleExportError(
                    f"detect_rules:{row_no} component_id={component_id!r} not found in rule_components "
                    f"for unit_id={unit_id}; known component_id values: {known}"
                )
        elif component_id is not None:
            warnings.append(
                f"detect_rules:{row_no} component_id is ignored for target={target} rule_id={rule_id}"
            )
        pattern = _required_text(row, "pattern", sheet="detect_rules", row_no=row_no)
        try:
            re.compile(pattern)
        except re.error as exc:
            raise BundleExportError(f"detect_rules:{row_no} regex compile failed for rule_id={rule_id}: {exc}") from exc
        if _looks_like_python_regex_literal(pattern):
            warnings.append(f"detect_rules:{row_no} pattern looks like a Python string literal for rule_id={rule_id}")

        hard_fail = _bool_value(row.get("hard_fail"), sheet="detect_rules", row_no=row_no, key="hard_fail")
        if stage == "verify" and not hard_fail:
            warnings.append(f"detect_rules:{row_no} verify rule has hard_fail=false for rule_id={rule_id}")

        rule = {
            "e_id": unit_id,
            "unit_id": unit_id,
            "source_e_id": _text(row.get("e_id")),
            "ps_id": _text(row.get("ps_id")),
            "ruleset_id": ruleset_id,
            "rule_id": rule_id,
            "stage": stage,
            "target": target,
            "pattern": pattern,
            "priority": _int_or_none(row.get("priority"), sheet="detect_rules", row_no=row_no, key="priority") or 0,
            "hard_fail": hard_fail,
            "component_id": component_id,
            "rule_type": "surface_regex",
            "engine": "re",
        }
        rules_by_ruleset_id[ruleset_id].append(rule)

    for rules in rules_by_ruleset_id.values():
        rules.sort(key=lambda item: (item["priority"], item["rule_id"]))

    all_ruleset_ids = set(rules_by_ruleset_id)
    for e_id, item in items_by_e_id.items():
        detect_ruleset_id = item.get("detect_ruleset_id")
        verify_ruleset_id = item.get("verify_ruleset_id")
        if detect_ruleset_id:
            _validate_ruleset_unit(
                owner_kind="items e_id",
                owner_id=e_id,
                ruleset_id=detect_ruleset_id,
                expected_unit_id=e_id,
                expected_stage="detect",
                rules_by_ruleset_id=rules_by_ruleset_id,
            )
        if verify_ruleset_id:
            _validate_ruleset_unit(
                owner_kind="items e_id",
                owner_id=e_id,
                ruleset_id=verify_ruleset_id,
                expected_unit_id=e_id,
                expected_stage="verify",
                rules_by_ruleset_id=rules_by_ruleset_id,
            )

    for ps_id, polyset in explicit_polysets_by_id.items():
        detect_ruleset_id = polyset.get("detect_ruleset_id")
        verify_ruleset_id = polyset.get("verify_ruleset_id")
        if detect_ruleset_id:
            _validate_ruleset_unit(
                owner_kind="polysets ps_id",
                owner_id=ps_id,
                ruleset_id=detect_ruleset_id,
                expected_unit_id=ps_id,
                expected_stage="detect",
                rules_by_ruleset_id=rules_by_ruleset_id,
            )
        if verify_ruleset_id:
            _validate_ruleset_unit(
                owner_kind="polysets ps_id",
                owner_id=ps_id,
                ruleset_id=verify_ruleset_id,
                expected_unit_id=ps_id,
                expected_stage="verify",
                rules_by_ruleset_id=rules_by_ruleset_id,
            )

    for ruleset_id, rules in rules_by_ruleset_id.items():
        stages = {rule["stage"] for rule in rules}
        if len(stages) > 1:
            raise BundleExportError(f"ruleset_id contains mixed stages: {ruleset_id} -> {sorted(stages)}")

    referenced_rulesets = {item.get("detect_ruleset_id") for item in items_by_e_id.values()} | {
        item.get("verify_ruleset_id") for item in items_by_e_id.values()
    } | {
        polyset.get("detect_ruleset_id") for polyset in explicit_polysets_by_id.values()
    } | {
        polyset.get("verify_ruleset_id") for polyset in explicit_polysets_by_id.values()
    }
    referenced_rulesets.discard(None)
    for ruleset_id in sorted(all_ruleset_ids - referenced_rulesets):
        warnings.append(f"ruleset_id not referenced by items/polysets: {ruleset_id}")

    polyset_members: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in items_by_e_id.values():
        polyset_id = item.get("ps_id") or item.get("polyset_id")
        if polyset_id:
            polyset_members[polyset_id].append(item)
    for polyset_id, polyset in explicit_polysets_by_id.items():
        if polyset_id not in polyset_members:
            polyset_members[polyset_id] = [items_by_e_id[e_id] for e_id in polyset["member_e_ids"]]

    polysets_by_id: dict[str, dict[str, Any]] = {}
    for polyset_id, members in sorted(polyset_members.items()):
        member_e_ids = [member["e_id"] for member in members]
        if polyset_id in explicit_polysets_by_id:
            polyset = dict(explicit_polysets_by_id[polyset_id])
            if polyset["member_e_ids"] != member_e_ids:
                declared = ";".join(polyset["member_e_ids"])
                discovered = ";".join(member_e_ids)
                warnings.append(
                    f"polysets ps_id={polyset_id} member order differs from items order: declared={declared}, items={discovered}"
                )
            polyset.setdefault("detect_form", polyset["canonical_form"])
        else:
            polyset = {
                "ps_id": polyset_id,
                "polyset_id": polyset_id,
                "primary_e_id": member_e_ids[0] if member_e_ids else None,
                "member_e_ids": member_e_ids,
                "canonical_form": _derive_polyset_form(members),
                "detect_form": _derive_polyset_form(members),
                "disconti_allowed": any(bool(member.get("disconti_allowed")) for member in members),
                "ps_comp_ids": sorted({comp_id for member in members for comp_id in member.get("e_comp_ids", [])}),
                "gloss_intro": None,
                "encoder_gloss": _join_encoder_gloss(None, members),
                "note": None,
                "detect_ruleset_id": None,
                "verify_ruleset_id": None,
            }
        polyset["detect_form"] = polyset.get("detect_form") or polyset.get("canonical_form")
        polysets_by_id[polyset_id] = polyset

    runtime_units: dict[str, dict[str, Any]] = {}
    for e_id, item in items_by_e_id.items():
        if item["group"] == "c" and (item.get("ps_id") or item.get("polyset_id")):
            continue
        detect_ruleset_ids = [item["detect_ruleset_id"]] if item.get("detect_ruleset_id") else []
        verify_ruleset_ids = [item["verify_ruleset_id"]] if item.get("verify_ruleset_id") else []
        runtime_units[e_id] = {
            "unit_id": e_id,
            "unit_type": "item",
            "group": item["group"],
            "member_e_ids": [e_id],
            "canonical_form": item["canonical_form"],
            "detect_ruleset_ids": detect_ruleset_ids,
            "verify_ruleset_ids": verify_ruleset_ids,
        }

    for polyset_id, polyset in polysets_by_id.items():
        members = [items_by_e_id[e_id] for e_id in polyset["member_e_ids"]]
        detect_ruleset_ids = [polyset["detect_ruleset_id"]] if polyset.get("detect_ruleset_id") else [
            member["detect_ruleset_id"] for member in members if member.get("detect_ruleset_id")
        ]
        verify_ruleset_ids = [polyset["verify_ruleset_id"]] if polyset.get("verify_ruleset_id") else [
            member["verify_ruleset_id"] for member in members if member.get("verify_ruleset_id")
        ]
        runtime_units[polyset_id] = {
            "unit_id": polyset_id,
            "unit_type": "polyset",
            "group": "c",
            "member_e_ids": polyset["member_e_ids"],
            "canonical_form": polyset.get("canonical_form") or polyset["detect_form"],
            "gloss": polyset.get("encoder_gloss"),
            "detect_ruleset_ids": detect_ruleset_ids,
            "verify_ruleset_ids": verify_ruleset_ids,
        }

    return {
        "schema_version": SCHEMA_VERSION,
        "source": {
            "dict_xlsx": str(dict_xlsx),
            "source_sha256": _sha256(dict_xlsx),
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "items_by_e_id": dict(sorted(items_by_e_id.items())),
        "components_by_e_id": {e_id: components_by_e_id[e_id] for e_id in sorted(components_by_e_id)},
        "bridges_by_id": {
            bridge_id: bridge_metadata_by_id()[bridge_id]
            for bridge_id in sorted(referenced_bridge_ids)
        },
        "rules_by_ruleset_id": {ruleset_id: rules_by_ruleset_id[ruleset_id] for ruleset_id in sorted(rules_by_ruleset_id)},
        "polysets_by_id": polysets_by_id,
        "runtime_units": dict(sorted(runtime_units.items())),
        "warnings": warnings,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export HanTalk detector runtime bundle from dict.xlsx.")
    parser.add_argument("--dict", dest="dict_xlsx", type=Path, default=Path("datasets/dict/dict.xlsx"))
    parser.add_argument("--out", type=Path, default=Path("configs/detector/detector_bundle.json"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        bundle = build_bundle(args.dict_xlsx)
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(json.dumps(bundle, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    except Exception as exc:  # noqa: BLE001 - CLI should give a concise fatal message.
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 2

    print(f"bundle={args.out}")
    print(f"schema_version={bundle['schema_version']}")
    print(f"items={len(bundle['items_by_e_id'])}")
    print(f"runtime_units={len(bundle['runtime_units'])}")
    print(f"warnings={len(bundle['warnings'])}")
    for warning in bundle["warnings"]:
        print(f"[WARNING] {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
