#!/usr/bin/env python3
"""Run sharded full-corpus offline review search for sparse-TP units.

This runner is used after sampled batch review indicates that TP examples are
too sparse for ordinary batch repetition. It never loads the full corpus into
memory. Instead, it prepares/searches 10,200-row shards using the top-level
example-making sampling ratio, applies Codex first-pass review, and stops when
the advisory TP count reaches a target.

Codex first-pass labels are advisory only. Human labels remain final.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover - handled when xlsx output is requested.
    openpyxl = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = None  # type: ignore[assignment]

try:
    from .apply_first_pass_review import apply_first_pass_review
    from .detector.engine import DetectorEngine
    from .prepare_codex_review import prepare_codex_review
    from .prepare_example_corpus import prepare_corpus_from_domain_plan
    from .search_corpus import search_corpus
    from .test_gold import evaluate_detector_bundle
    from .validate_dict_bundle_sync import validate_dict_bundle_sync
except ImportError:  # pragma: no cover - supports direct script execution.
    from apply_first_pass_review import apply_first_pass_review
    from detector.engine import DetectorEngine
    from prepare_codex_review import prepare_codex_review
    from prepare_example_corpus import prepare_corpus_from_domain_plan
    from search_corpus import search_corpus
    from test_gold import evaluate_detector_bundle
    from validate_dict_bundle_sync import validate_dict_bundle_sync


SCHEMA_VERSION = "hantalk_full_corpus_review_run_v1"
SHARD_SCHEMA_VERSION = "hantalk_full_corpus_review_shard_v1"
GOLD_BUNDLE_MATCH_POLICY = "overlap"
DEFAULT_TARGET_FIRST_PASS_TP = 150
DEFAULT_MAX_SHARDS = 50
DEFAULT_BACKFILL_DOMAIN = "news"
LABEL_KEYS = ("tp", "fp", "unclear", "blank")


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_json(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")


def _load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_manifest_sampling(manifest_path: Path) -> dict[str, int]:
    data = _load_json(manifest_path)
    sampling = data.get("sampling") or {}
    if not sampling:
        raise ValueError(f"{manifest_path}: missing top-level sampling")
    return {str(domain): int(size) for domain, size in sampling.items()}


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                records.append(json.loads(stripped))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL at {path}:{line_no}: {exc}") from exc
    return records


def _ensure_file(path: Path, *, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} not found: {path}")
    if not path.is_file():
        raise ValueError(f"{label} is not a file: {path}")


def _csv_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
    return str(value)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"{path}: missing header row")
        columns = [str(name or "").strip() for name in reader.fieldnames]
        if any(not column for column in columns):
            raise ValueError(f"{path}: blank header")
        if len(set(columns)) != len(columns):
            duplicates = sorted(column for column in set(columns) if columns.count(column) > 1)
            raise ValueError(f"{path}: duplicate header after strip: {duplicates}")
        rows = [
            {
                normalized: str(row.get(original) or "").strip()
                for original, normalized in zip(reader.fieldnames, columns)
            }
            for row in reader
        ]
    return columns, rows


def _write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_cell(row.get(column, "")) for column in columns})


def _write_xlsx(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to write .xlsx files")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "full_corpus_first_pass"
    sheet.append(columns)
    for row in rows:
        sheet.append([_csv_cell(row.get(column, "")) for column in columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    tp_fill = PatternFill("solid", fgColor="E2F0D9")
    fp_fill = PatternFill("solid", fgColor="FCE4D6")
    unclear_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_index = {name: idx + 1 for idx, name in enumerate(columns)}
    label_col = column_index.get("codex_review_label")
    for row_idx in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row_idx, label_col).value or "") if label_col else ""
        fill = tp_fill if label == "tp" else fp_fill if label == "fp" else unclear_fill if label == "unclear" else None
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill is not None:
                cell.fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    width_overrides = {
        "raw_text": 70,
        "regex_with_context": 70,
        "regex_match_text": 18,
        "a_token_left": 18,
        "two_tokens_right": 24,
        "codex_review_label": 18,
        "codex_review_span_status": 22,
        "codex_review_reason": 32,
        "codex_review_note": 52,
        "full_corpus_shard_index": 18,
        "full_corpus_run_id": 24,
        "memo": 36,
    }
    for idx, column in enumerate(columns, start=1):
        width = width_overrides.get(column, min(max(len(column) + 2, 10), 24))
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    workbook.save(path)


def _label_counts(raw_counts: dict[str, Any] | None) -> dict[str, int]:
    raw_counts = raw_counts or {}
    return {key: int(raw_counts.get(key) or 0) for key in LABEL_KEYS}


def _accumulate_shard_counts(cumulative: Counter[str], shard: dict[str, Any]) -> None:
    search_summary = shard.get("search_summary") or {}
    label_counts = _label_counts(shard.get("first_pass_label_counts"))
    cumulative["n_input_texts"] += int(search_summary.get("n_input_texts") or 0)
    cumulative["n_candidates"] += int(search_summary.get("n_candidates") or 0)
    cumulative["first_pass_tp"] += label_counts["tp"]
    cumulative["first_pass_fp"] += label_counts["fp"]
    cumulative["first_pass_unclear"] += label_counts["unclear"]
    cumulative["first_pass_blank"] += label_counts["blank"]


def _initial_domain_state(base_quota: dict[str, int]) -> dict[str, dict[str, Any]]:
    return {
        domain: {
            "rank_cursor": 0,
            "exhausted": False,
            "exhausted_reason": None,
        }
        for domain in base_quota
    }


def _copy_domain_state(state: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        domain: {
            "rank_cursor": int(values.get("rank_cursor") or 0),
            "exhausted": bool(values.get("exhausted")),
            "exhausted_reason": values.get("exhausted_reason"),
        }
        for domain, values in state.items()
    }


def _all_domains_exhausted(state: dict[str, dict[str, Any]]) -> bool:
    return all(bool(values.get("exhausted")) for values in state.values())


def _build_domain_plan(
    *,
    base_quota: dict[str, int],
    state: dict[str, dict[str, Any]],
    backfill_domain: str,
) -> tuple[dict[str, dict[str, int]], dict[str, Any]]:
    if backfill_domain not in base_quota:
        raise ValueError(f"backfill domain is not in manifest sampling: {backfill_domain}")
    requested_by_domain = {domain: 0 for domain in base_quota}
    shortfall_by_domain = {domain: 0 for domain in base_quota}
    backfill_added = 0

    for domain, quota in base_quota.items():
        if domain == backfill_domain:
            continue
        if bool(state[domain].get("exhausted")):
            shortfall_by_domain[domain] = int(quota)
            backfill_added += int(quota)
        else:
            requested_by_domain[domain] = int(quota)

    if bool(state[backfill_domain].get("exhausted")):
        shortfall_by_domain[backfill_domain] = int(base_quota[backfill_domain])
        requested_by_domain[backfill_domain] = 0
    else:
        requested_by_domain[backfill_domain] = int(base_quota[backfill_domain]) + backfill_added

    domain_plan = {
        domain: {
            "requested": int(requested),
            "rank_start": int(state[domain].get("rank_cursor") or 0),
        }
        for domain, requested in requested_by_domain.items()
    }
    return domain_plan, {
        "requested_by_domain": requested_by_domain,
        "shortfall_by_domain": shortfall_by_domain,
        "backfill_added_to_domain": backfill_added,
        "backfill_domain": backfill_domain,
    }


def _update_domain_state_after_prepare(
    *,
    state_before: dict[str, dict[str, Any]],
    requested_by_domain: dict[str, int],
    selected_by_domain: dict[str, int],
) -> tuple[dict[str, dict[str, Any]], dict[str, int]]:
    state_after = _copy_domain_state(state_before)
    shortfall_by_domain: dict[str, int] = {}
    for domain, requested in requested_by_domain.items():
        requested = int(requested or 0)
        selected = int(selected_by_domain.get(domain) or 0)
        state_after[domain]["rank_cursor"] = int(state_after[domain].get("rank_cursor") or 0) + selected
        shortfall = max(0, requested - selected)
        shortfall_by_domain[domain] = shortfall
        if requested > 0 and selected < requested:
            state_after[domain]["exhausted"] = True
            state_after[domain]["exhausted_reason"] = "selected_below_requested"
    return state_after, shortfall_by_domain


def _output_root(*, artifact_root: Path, unit_id: str) -> Path:
    return artifact_root / unit_id / "full_corpus"


def _batch_id(*, unit_id: str, shard_index: int) -> str:
    return f"full_corpus_{unit_id}_shard_{shard_index:03d}"


def _prepared_paths(*, prepared_root: Path, unit_id: str, batch_id: str) -> dict[str, Path]:
    item_dir = prepared_root / unit_id
    return {
        "prepared_jsonl": item_dir / f"{batch_id}.jsonl",
        "prepared_report_json": item_dir / f"{batch_id}_report.json",
    }


def _shard_paths(*, artifact_root: Path, unit_id: str, shard_index: int) -> dict[str, Path]:
    root = _output_root(artifact_root=artifact_root, unit_id=unit_id)
    label = f"{unit_id}_full_shard_{shard_index:03d}"
    return {
        "root": root,
        "shard_report_json": root / f"{label}_report.json",
        "detection_jsonl": root / f"{label}_detection.jsonl",
        "human_review_csv": root / f"{label}_human_review.csv",
        "search_report_json": root / f"{label}_search_report.json",
        "codex_review_csv": root / f"{label}_codex_review.csv",
        "codex_review_xlsx": root / f"{label}_codex_review.xlsx",
        "codex_review_report_json": root / f"{label}_codex_review_report.json",
        "first_pass_csv": root / f"{label}_codex_review_first_pass.csv",
        "first_pass_xlsx": root / f"{label}_codex_review_first_pass.xlsx",
        "first_pass_report_json": root / f"{label}_codex_review_first_pass_report.json",
    }


def _final_paths(*, artifact_root: Path, unit_id: str) -> dict[str, Path]:
    root = _output_root(artifact_root=artifact_root, unit_id=unit_id)
    return {
        "root": root,
        "merged_first_pass_csv": root / f"{unit_id}_full_corpus_first_pass_merged.csv",
        "merged_first_pass_xlsx": root / f"{unit_id}_full_corpus_first_pass_merged.xlsx",
        "run_report_json": root / f"{unit_id}_full_corpus_run_report.json",
    }


def _remove_shard_outputs(paths: dict[str, Path]) -> None:
    for key, path in paths.items():
        if key == "root":
            continue
        if path.exists():
            if not path.is_file():
                raise FileExistsError(f"Cannot overwrite non-file path: {path}")
            path.unlink()


def _gold_gate(*, unit_id: str, gold_path: Path, bundle_path: Path, allow_polyset: bool) -> dict[str, Any]:
    gold_records = _read_jsonl(gold_path)
    engine = DetectorEngine.from_bundle(bundle_path)
    result = evaluate_detector_bundle(
        item_id=unit_id,
        engine=engine,
        gold_records=gold_records,
        active_unit_ids=[unit_id],
        bundle_match_policy=GOLD_BUNDLE_MATCH_POLICY,
        allow_polyset=allow_polyset,
    )
    fn_records = result.pop("fn_records", [])
    result["fn_records_preview"] = fn_records[:20]
    result["fn_records_preview_truncated"] = len(fn_records) > 20
    return result


def _gold_gate_ok(result: dict[str, Any]) -> bool:
    return float(result.get("gold_recall") or 0.0) >= 1.0 and int(result.get("fn_count") or 0) == 0


def _validate_existing_shard(shard_report: dict[str, Any], *, unit_id: str, shard_index: int) -> None:
    if shard_report.get("schema_version") != SHARD_SCHEMA_VERSION:
        raise ValueError(
            f"Existing shard report schema mismatch: {shard_report.get('schema_version')!r} != {SHARD_SCHEMA_VERSION!r}"
        )
    if str(shard_report.get("unit_id") or "") != unit_id:
        raise ValueError(f"Existing shard unit_id mismatch: {shard_report.get('unit_id')!r} != {unit_id!r}")
    report_shard_index = shard_report.get("shard_index")
    if report_shard_index is None or int(report_shard_index) != shard_index:
        raise ValueError(f"Existing shard_index mismatch: {shard_report.get('shard_index')!r} != {shard_index!r}")
    status = str(shard_report.get("status") or "")
    if status not in {"ok", "ok_no_candidates"}:
        raise ValueError(f"Existing shard report is not reusable: status={status!r}")
    sampling = shard_report.get("sampling") or {}
    if not sampling.get("domain_state_after"):
        raise ValueError("Existing shard report is missing sampling.domain_state_after")
    if not shard_report.get("first_pass_label_counts"):
        raise ValueError("Existing shard report is missing first_pass_label_counts")
    outputs = shard_report.get("outputs") or {}
    n_candidates = int((shard_report.get("search_summary") or {}).get("n_candidates") or 0)
    required_keys = ["prepared_jsonl", "prepared_report_json", "detection_jsonl", "search_report_json"]
    if n_candidates > 0:
        required_keys.extend(
            [
                "human_review_csv",
                "codex_review_csv",
                "codex_review_xlsx",
                "first_pass_csv",
                "first_pass_xlsx",
                "first_pass_report_json",
            ]
        )
    missing = [key for key in required_keys if not outputs.get(key) or not Path(str(outputs[key])).exists()]
    if missing:
        raise FileNotFoundError(f"Existing shard artifact is incomplete. Missing output(s): {missing}")


def _reuse_existing_shard(shard_report_path: Path, *, unit_id: str, shard_index: int) -> dict[str, Any]:
    shard_report = _load_json(shard_report_path)
    _validate_existing_shard(shard_report, unit_id=unit_id, shard_index=shard_index)
    reused = dict(shard_report)
    reused["status"] = str(shard_report.get("status") or "ok")
    reused["run_step_status"] = "skipped_existing_report"
    return reused


def _prepared_report_matches_domain_plan(prepare_report: dict[str, Any], domain_plan: dict[str, dict[str, int]]) -> bool:
    requested = {
        str(domain): int(size)
        for domain, size in (prepare_report.get("sampling_requested_by_domain") or {}).items()
    }
    rank_start = {
        str(domain): int(offset)
        for domain, offset in (prepare_report.get("sampling_rank_start_by_domain") or {}).items()
    }
    expected_requested = {domain: int(plan.get("requested") or 0) for domain, plan in domain_plan.items()}
    expected_rank_start = {domain: int(plan.get("rank_start") or 0) for domain, plan in domain_plan.items()}
    return requested == expected_requested and rank_start == expected_rank_start


def _run_one_shard(
    *,
    unit_id: str,
    shard_index: int,
    args: argparse.Namespace,
    prepared: dict[str, Path],
    outputs: dict[str, Path],
    batch_id: str,
    domain_plan: dict[str, dict[str, int]],
    domain_state_before: dict[str, dict[str, Any]],
    sampling_plan_summary: dict[str, Any],
) -> dict[str, Any]:
    outputs["root"].mkdir(parents=True, exist_ok=True)
    prepared["prepared_jsonl"].parent.mkdir(parents=True, exist_ok=True)

    if args.overwrite_shards:
        _remove_shard_outputs(outputs)
        for prepared_path in prepared.values():
            if prepared_path.exists():
                if not prepared_path.is_file():
                    raise FileExistsError(f"Cannot overwrite non-file prepared path: {prepared_path}")
                prepared_path.unlink()

    shard_report: dict[str, Any] = {
        "schema_version": SHARD_SCHEMA_VERSION,
        "unit_id": unit_id,
        "shard_index": shard_index,
        "batch_id": batch_id,
        "created_at": _now_utc(),
        "status": "running",
        "run_step_status": "running",
        "first_pass_label_counts": _label_counts({}),
        "prepare_corpus_summary": None,
        "search_summary": None,
        "codex_review_summary": None,
        "first_pass_review_summary": None,
        "sampling": {
            "domain_plan": domain_plan,
            "requested_by_domain": sampling_plan_summary["requested_by_domain"],
            "selected_by_domain": {},
            "shortfall_by_domain": sampling_plan_summary["shortfall_by_domain"],
            "backfill_domain": sampling_plan_summary["backfill_domain"],
            "backfill_added_to_domain": sampling_plan_summary["backfill_added_to_domain"],
            "domain_state_before": _copy_domain_state(domain_state_before),
            "domain_state_after": _copy_domain_state(domain_state_before),
        },
        "outputs": {
            "prepared_jsonl": str(prepared["prepared_jsonl"]),
            "prepared_report_json": str(prepared["prepared_report_json"]),
            "detection_jsonl": str(outputs["detection_jsonl"]),
            "human_review_csv": str(outputs["human_review_csv"]),
            "search_report_json": str(outputs["search_report_json"]),
            "codex_review_csv": str(outputs["codex_review_csv"]),
            "codex_review_xlsx": str(outputs["codex_review_xlsx"]),
            "codex_review_report_json": str(outputs["codex_review_report_json"]),
            "first_pass_csv": str(outputs["first_pass_csv"]),
            "first_pass_xlsx": str(outputs["first_pass_xlsx"]),
            "first_pass_report_json": str(outputs["first_pass_report_json"]),
            "shard_report_json": str(outputs["shard_report_json"]),
        },
    }

    try:
        if prepared["prepared_jsonl"].exists() and prepared["prepared_report_json"].exists():
            prepare_report = _load_json(prepared["prepared_report_json"])
            if not _prepared_report_matches_domain_plan(prepare_report, domain_plan):
                raise ValueError(
                    "Existing prepared shard does not match requested domain plan. "
                    "Use --overwrite-shards to regenerate it."
                )
            shard_report["prepare_corpus_summary"] = prepare_report
            shard_report["prepare_corpus_step_status"] = "skipped_existing"
        else:
            prepare_report = prepare_corpus_from_domain_plan(
                manifest_path=args.manifest,
                corpus_root=args.corpus_root,
                batch_id=batch_id,
                shard_index=shard_index,
                domain_plan=domain_plan,
                out_path=prepared["prepared_jsonl"],
                report_path=prepared["prepared_report_json"],
                seed_override=args.seed,
            )
            shard_report["prepare_corpus_summary"] = prepare_report
            shard_report["prepare_corpus_step_status"] = "ok"
        selected_by_domain = {
            domain: int(count) for domain, count in (prepare_report.get("n_rows_selected_by_domain") or {}).items()
        }
        domain_state_after, actual_shortfall_by_domain = _update_domain_state_after_prepare(
            state_before=domain_state_before,
            requested_by_domain=sampling_plan_summary["requested_by_domain"],
            selected_by_domain=selected_by_domain,
        )
        shard_report["sampling"]["selected_by_domain"] = selected_by_domain
        shard_report["sampling"]["shortfall_by_domain"] = actual_shortfall_by_domain
        shard_report["sampling"]["domain_state_after"] = domain_state_after

        search_report = search_corpus(
            bundle_path=args.bundle,
            input_jsonl=prepared["prepared_jsonl"],
            active_unit_ids=[unit_id],
            out_jsonl=outputs["detection_jsonl"],
            review_csv=outputs["human_review_csv"],
            report_json=outputs["search_report_json"],
            allow_polyset=args.allow_polyset,
            include_debug=args.include_debug,
        )
        n_candidates = int(search_report.get("n_candidates") or 0)
        shard_report["search_summary"] = {
            "n_input_texts": search_report.get("n_input_texts"),
            "n_input_by_domain": search_report.get("n_input_by_domain"),
            "n_texts_with_hits": search_report.get("n_texts_with_hits"),
            "n_candidates": n_candidates,
            "n_candidates_by_domain": search_report.get("n_candidates_by_domain"),
            "n_candidates_before_verify": search_report.get("n_candidates_before_verify"),
            "n_candidates_after_verify": search_report.get("n_candidates_after_verify"),
            "n_candidates_hard_failed": search_report.get("n_candidates_hard_failed"),
            "span_source_counts": search_report.get("span_source_counts"),
            "component_span_status_counts": search_report.get("component_span_status_counts"),
            "elapsed_sec": search_report.get("elapsed_sec"),
        }

        if n_candidates == 0:
            shard_report["status"] = "ok_no_candidates"
            shard_report["run_step_status"] = "ok_no_candidates"
            return shard_report

        codex_report = prepare_codex_review(
            item_id=unit_id,
            input_path=outputs["human_review_csv"],
            out_csv=outputs["codex_review_csv"],
            out_xlsx=outputs["codex_review_xlsx"],
            report_json=outputs["codex_review_report_json"],
        )
        shard_report["codex_review_summary"] = {
            "n_rows": codex_report.get("n_rows"),
            "span_parse_counts": codex_report.get("span_parse_counts"),
            "span_source_counts": codex_report.get("span_source_counts"),
            "component_span_status_counts": codex_report.get("component_span_status_counts"),
        }
        first_pass_report = apply_first_pass_review(
            item_id=unit_id,
            input_path=outputs["codex_review_csv"],
            out_csv=outputs["first_pass_csv"],
            out_xlsx=outputs["first_pass_xlsx"],
            report_json=outputs["first_pass_report_json"],
        )
        label_counts = _label_counts(first_pass_report.get("codex_review_label_counts"))
        shard_report["first_pass_label_counts"] = label_counts
        shard_report["first_pass_review_summary"] = {
            "profile_id": first_pass_report.get("profile_id"),
            "profile_status": first_pass_report.get("profile_status"),
            "profile_source": first_pass_report.get("profile_source"),
            "advisory_labels_applied": first_pass_report.get("advisory_labels_applied"),
            "n_rows": first_pass_report.get("n_rows"),
            "codex_review_label_counts": first_pass_report.get("codex_review_label_counts"),
            "codex_review_reason_counts": first_pass_report.get("codex_review_reason_counts"),
            "codex_review_reason_ko_counts": first_pass_report.get("codex_review_reason_ko_counts"),
        }
        shard_report["status"] = "ok"
        shard_report["run_step_status"] = "ok"
        return shard_report
    finally:
        shard_report["finished_at"] = _now_utc()
        _write_json(outputs["shard_report_json"], shard_report)


def _label_sort_value(value: str) -> int:
    order = {"tp": 0, "unclear": 1, "fp": 2, "blank": 3, "": 4}
    return order.get(value, 5)


def _merge_first_pass_files(
    *,
    unit_id: str,
    run_id: str,
    shards: list[dict[str, Any]],
    out_csv: Path,
    out_xlsx: Path,
) -> dict[str, Any]:
    rows: list[dict[str, str]] = []
    columns: list[str] = []
    seen_columns: set[str] = set()
    for shard in shards:
        outputs = shard.get("outputs") or {}
        first_pass_csv = outputs.get("first_pass_csv")
        if not first_pass_csv:
            continue
        path = Path(str(first_pass_csv))
        if not path.exists():
            continue
        input_columns, input_rows = _read_csv(path)
        for column in ["full_corpus_run_id", "full_corpus_shard_index"]:
            if column not in seen_columns:
                columns.append(column)
                seen_columns.add(column)
        for column in input_columns:
            if column not in seen_columns:
                columns.append(column)
                seen_columns.add(column)
        shard_index = str(shard.get("shard_index"))
        for row in input_rows:
            merged = dict(row)
            merged["full_corpus_run_id"] = run_id
            merged["full_corpus_shard_index"] = shard_index
            rows.append(merged)

    if not columns:
        columns = ["full_corpus_run_id", "full_corpus_shard_index", "unit_id", "hit_id", "raw_text", "codex_review_label"]
    rows.sort(
        key=lambda row: (
            _label_sort_value(str(row.get("codex_review_label") or "")),
            str(row.get("codex_review_reason") or ""),
            str(row.get("corpus_domain") or ""),
            str(row.get("regex_match_text") or ""),
            str(row.get("hit_id") or ""),
        )
    )
    _write_csv(out_csv, rows, columns)
    _write_xlsx(out_xlsx, rows, columns)
    label_counts = Counter(str(row.get("codex_review_label") or "blank").strip() or "blank" for row in rows)
    return {
        "unit_id": unit_id,
        "n_rows": len(rows),
        "codex_review_label_counts": _label_counts(dict(label_counts)),
        "out_csv": str(out_csv),
        "out_xlsx": str(out_xlsx),
    }


def run_full_corpus_review(args: argparse.Namespace) -> dict[str, Any]:
    unit_id = str(args.unit_id).strip()
    if not unit_id:
        raise ValueError("--unit-id must not be blank")
    if args.start_shard_index < 0:
        raise ValueError("--start-shard-index must be >= 0")
    if args.max_shards <= 0:
        raise ValueError("--max-shards must be > 0")
    if args.target_first_pass_tp <= 0:
        raise ValueError("--target-first-pass-tp must be > 0")
    for path, label in [
        (args.gold, "gold JSONL"),
        (args.bundle, "detector bundle"),
        (args.manifest, "corpus manifest"),
    ]:
        _ensure_file(path, label=label)
    if args.dict_xlsx is not None:
        _ensure_file(args.dict_xlsx, label="dict Excel")
    if not args.corpus_root.exists():
        raise FileNotFoundError(f"corpus root not found: {args.corpus_root}")
    base_quota = _load_manifest_sampling(args.manifest)
    if args.backfill_domain not in base_quota:
        raise ValueError(f"--backfill-domain is not in manifest sampling: {args.backfill_domain}")

    final = _final_paths(artifact_root=args.artifact_root, unit_id=unit_id)
    final["root"].mkdir(parents=True, exist_ok=True)
    run_id = f"{unit_id}_full_corpus"
    report: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "unit_id": unit_id,
        "run_id": run_id,
        "created_at": _now_utc(),
        "mode": "full_corpus_offline_sharded",
        "target_first_pass_tp": args.target_first_pass_tp,
        "target_basis": "codex_first_pass_advisory_label_not_human_final",
        "max_shards": args.max_shards,
        "start_shard_index": args.start_shard_index,
        "inputs": {
            "gold": str(args.gold),
            "bundle": str(args.bundle),
            "dict": str(args.dict_xlsx) if args.dict_xlsx else None,
            "manifest": str(args.manifest),
            "corpus_root": str(args.corpus_root),
            "prepared_root": str(args.prepared_root),
            "artifact_root": str(args.artifact_root),
            "allow_polyset": bool(args.allow_polyset),
            "seed": args.seed,
        },
        "sampling_policy": {
            "base_quota": base_quota,
            "backfill_domain": args.backfill_domain,
            "target_rows_per_shard": int(sum(base_quota.values())),
            "cursor_advances_by": "selected_count",
            "selected_below_requested_means_exhausted": True,
            "backfill_chain": False,
            "stop_when_backfill_domain_exhausted": True,
        },
        "gold_gate": None,
        "dict_bundle_sync": None,
        "shards": [],
        "cumulative": {
            "n_input_texts": 0,
            "n_candidates": 0,
            "first_pass_tp": 0,
            "first_pass_fp": 0,
            "first_pass_unclear": 0,
            "first_pass_blank": 0,
        },
        "shards_processed": 0,
        "shards_reused": 0,
        "stop_reason": None,
        "target_reached": False,
        "final_domain_state": None,
        "outputs": {
            "merged_first_pass_csv": str(final["merged_first_pass_csv"]),
            "merged_first_pass_xlsx": str(final["merged_first_pass_xlsx"]),
            "run_report_json": str(final["run_report_json"]),
        },
        "status": "running",
        "failure_reason": None,
    }

    try:
        gold_result = _gold_gate(
            unit_id=unit_id,
            gold_path=args.gold,
            bundle_path=args.bundle,
            allow_polyset=args.allow_polyset,
        )
        report["gold_gate"] = {
            "bundle_match_policy": GOLD_BUNDLE_MATCH_POLICY,
            "gold_total": gold_result.get("gold_total"),
            "gold_recall": gold_result.get("gold_recall"),
            "fn_count": gold_result.get("fn_count"),
            "span_exact_recall": gold_result.get("span_exact_recall"),
            "component_span_success_count": gold_result.get("component_span_success_count"),
            "fn_records_preview": gold_result.get("fn_records_preview"),
        }
        if not _gold_gate_ok(gold_result):
            report["status"] = "blocked"
            report["failure_reason"] = "gold_recall_fix_required"
            report["stop_reason"] = "gold_recall_fix_required"
            return report

        if args.dict_xlsx is not None and not args.skip_dict_bundle_sync:
            sync_report = validate_dict_bundle_sync(
                dict_xlsx=args.dict_xlsx,
                bundle_path=args.bundle,
                unit_id=unit_id,
                report_json=final["root"] / f"{unit_id}_full_corpus_dict_bundle_sync_report.json",
            )
            report["dict_bundle_sync"] = {
                "in_sync": sync_report.get("in_sync"),
                "diff_count": sync_report.get("diff_count"),
                "dict_export_warning_count": sync_report.get("dict_export_warning_count"),
                "report_json": str(final["root"] / f"{unit_id}_full_corpus_dict_bundle_sync_report.json"),
            }
            if not sync_report.get("in_sync"):
                report["status"] = "blocked"
                report["failure_reason"] = "dict_bundle_mismatch"
                report["stop_reason"] = "dict_bundle_mismatch"
                return report
        else:
            report["dict_bundle_sync"] = {
                "in_sync": None,
                "status": "skipped_no_dict" if args.dict_xlsx is None else "skipped_by_flag",
            }

        cumulative = Counter()
        shards: list[dict[str, Any]] = []
        domain_state = _initial_domain_state(base_quota)
        if args.start_shard_index > 0:
            for prior_shard_index in range(0, args.start_shard_index):
                prior_outputs = _shard_paths(
                    artifact_root=args.artifact_root,
                    unit_id=unit_id,
                    shard_index=prior_shard_index,
                )
                if not prior_outputs["shard_report_json"].exists():
                    raise FileNotFoundError(
                        "Cannot start from a later shard without reusable prior shard reports. "
                        f"Missing: {prior_outputs['shard_report_json']}"
                    )
                prior_shard = _reuse_existing_shard(
                    prior_outputs["shard_report_json"],
                    unit_id=unit_id,
                    shard_index=prior_shard_index,
                )
                shards.append(prior_shard)
                report["shards_reused"] += 1
                sampling = prior_shard.get("sampling") or {}
                domain_state = _copy_domain_state(sampling.get("domain_state_after") or domain_state)
                _accumulate_shard_counts(cumulative, prior_shard)

            if cumulative["first_pass_tp"] >= args.target_first_pass_tp:
                report["stop_reason"] = "target_first_pass_tp_reached"
                report["target_reached"] = True

        for shard_index in range(args.start_shard_index, args.start_shard_index + args.max_shards):
            if report["stop_reason"] is not None:
                break
            batch_id = _batch_id(unit_id=unit_id, shard_index=shard_index)
            prepared = _prepared_paths(prepared_root=args.prepared_root, unit_id=unit_id, batch_id=batch_id)
            outputs = _shard_paths(artifact_root=args.artifact_root, unit_id=unit_id, shard_index=shard_index)
            if outputs["shard_report_json"].exists() and not args.overwrite_shards:
                shard = _reuse_existing_shard(
                    outputs["shard_report_json"],
                    unit_id=unit_id,
                    shard_index=shard_index,
                )
                report["shards_reused"] += 1
            else:
                domain_plan, sampling_plan_summary = _build_domain_plan(
                    base_quota=base_quota,
                    state=domain_state,
                    backfill_domain=args.backfill_domain,
                )
                shard = _run_one_shard(
                    unit_id=unit_id,
                    shard_index=shard_index,
                    args=args,
                    prepared=prepared,
                    outputs=outputs,
                    batch_id=batch_id,
                    domain_plan=domain_plan,
                    domain_state_before=domain_state,
                    sampling_plan_summary=sampling_plan_summary,
                )
            shards.append(shard)
            sampling = shard.get("sampling") or {}
            domain_state = _copy_domain_state(sampling.get("domain_state_after") or domain_state)
            _accumulate_shard_counts(cumulative, shard)

            if cumulative["first_pass_tp"] >= args.target_first_pass_tp:
                report["stop_reason"] = "target_first_pass_tp_reached"
                report["target_reached"] = True
                break
            if _all_domains_exhausted(domain_state):
                report["stop_reason"] = "all_corpora_exhausted_before_target"
                report["target_reached"] = False
                break
            if bool(domain_state.get(args.backfill_domain, {}).get("exhausted")):
                report["stop_reason"] = "backfill_domain_exhausted_before_target"
                report["target_reached"] = False
                break

        if report["stop_reason"] is None:
            report["stop_reason"] = "max_shards_reached"
            report["target_reached"] = False

        merge_report = _merge_first_pass_files(
            unit_id=unit_id,
            run_id=run_id,
            shards=shards,
            out_csv=final["merged_first_pass_csv"],
            out_xlsx=final["merged_first_pass_xlsx"],
        )
        report["shards"] = [
            {
                "shard_index": shard.get("shard_index"),
                "batch_id": shard.get("batch_id"),
                "status": shard.get("status"),
                "run_step_status": shard.get("run_step_status"),
                "search_summary": shard.get("search_summary"),
                "first_pass_label_counts": _label_counts(shard.get("first_pass_label_counts")),
                "sampling": shard.get("sampling"),
                "outputs": shard.get("outputs"),
            }
            for shard in shards
        ]
        report["cumulative"] = {key: int(value) for key, value in cumulative.items()}
        report["shards_processed"] = len(shards)
        report["merge_summary"] = merge_report
        report["final_domain_state"] = domain_state
        report["status"] = "ok"
        report["failure_reason"] = None
        if report["stop_reason"] == "max_shards_reached":
            report["next_step_hint"] = "Inspect merged first-pass file or rerun with a larger --max-shards value."
        elif report["stop_reason"] == "backfill_domain_exhausted_before_target":
            report["next_step_hint"] = "Backfill domain was exhausted before advisory TP target. Review the merged first-pass file with currently found examples."
        elif report["stop_reason"] == "all_corpora_exhausted_before_target":
            report["next_step_hint"] = "All corpora were exhausted before advisory TP target. Review the merged first-pass file with currently found examples."
        else:
            report["next_step_hint"] = "Review merged first-pass file and write final human_label/span_status."
        return report
    except Exception as exc:
        report["status"] = "failed"
        report["failure_reason"] = f"{type(exc).__name__}: {exc}"
        report["stop_reason"] = str(report.get("stop_reason") or "failed")
        raise
    finally:
        report["finished_at"] = _now_utc()
        _write_json(final["run_report_json"], report)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--unit-id", required=True, help="Detector runtime unit id, e.g. ps_df004.")
    parser.add_argument("--gold", required=True, type=Path, help="Exported gold JSONL path.")
    parser.add_argument("--bundle", required=True, type=Path, help="Detector bundle JSON path.")
    parser.add_argument("--dict", dest="dict_xlsx", type=Path, default=None, help="Optional dict Excel for sync validation.")
    parser.add_argument("--manifest", required=True, type=Path, help="Corpus manifest JSON.")
    parser.add_argument("--corpus-root", required=True, type=Path, help="Root folder containing corpus source files.")
    parser.add_argument("--prepared-root", required=True, type=Path, help="Root folder for full-corpus prepared shards.")
    parser.add_argument("--artifact-root", required=True, type=Path, help="Artifact root for unit-specific outputs.")
    parser.add_argument("--start-shard-index", type=int, default=0, help="Full-corpus shard index to start from.")
    parser.add_argument("--target-first-pass-tp", type=int, default=DEFAULT_TARGET_FIRST_PASS_TP)
    parser.add_argument("--max-shards", type=int, default=DEFAULT_MAX_SHARDS)
    parser.add_argument(
        "--backfill-domain",
        default=DEFAULT_BACKFILL_DOMAIN,
        help="Domain that receives shortfall when another domain is exhausted. Default: news.",
    )
    parser.add_argument("--allow-polyset", action="store_true", help="Allow multi-member ps_id polyset runtime units.")
    parser.add_argument("--include-debug", action="store_true", help="Include detector debug details in search output.")
    parser.add_argument("--overwrite-shards", action="store_true", help="Regenerate existing shard artifacts instead of reusing shard reports.")
    parser.add_argument("--skip-dict-bundle-sync", action="store_true", help="Skip dict/bundle sync validation when --dict is provided.")
    parser.add_argument("--seed", type=int, default=None, help="Override manifest seed.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)
    try:
        report = run_full_corpus_review(args)
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1

    print(
        json.dumps(
            {
                "unit_id": report["unit_id"],
                "status": report["status"],
                "failure_reason": report.get("failure_reason"),
                "stop_reason": report.get("stop_reason"),
                "target_first_pass_tp": report.get("target_first_pass_tp"),
                "target_reached": report.get("target_reached"),
                "cumulative": report.get("cumulative"),
                "shards_processed": report.get("shards_processed"),
                "shards_reused": report.get("shards_reused"),
                "merged_first_pass_xlsx": (report.get("outputs") or {}).get("merged_first_pass_xlsx"),
                "run_report_json": (report.get("outputs") or {}).get("run_report_json"),
                "next_step_hint": report.get("next_step_hint"),
            },
            ensure_ascii=False,
            indent=2,
            allow_nan=False,
        )
    )
    return 0 if report["status"] == "ok" else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
