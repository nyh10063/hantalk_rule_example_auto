#!/usr/bin/env python3
"""Export human-labeled review rows into HanTalk encoder pair examples."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - handled at runtime for xlsx I/O.
    openpyxl = None  # type: ignore[assignment]

try:
    from .detector.span_utils import (
        DEFAULT_GAP_MARKER,
        format_span_segments,
        inject_span_markers,
        make_span_key,
        make_span_text,
        parse_span_segments,
        validate_span_segments,
    )
except ImportError:  # pragma: no cover - supports direct script execution.
    from detector.span_utils import (
        DEFAULT_GAP_MARKER,
        format_span_segments,
        inject_span_markers,
        make_span_key,
        make_span_text,
        parse_span_segments,
        validate_span_segments,
    )

SCHEMA_VERSION = "hantalk_encoder_pair_example_v1"
SUMMARY_SCHEMA_VERSION = "hantalk_encoder_examples_export_summary_v1"
INPUT_CONSTRUCTION_VERSION = "hantalk_binary_pair_v1"
SPAN_MARKER_STYLE = "[SPAN]...[/SPAN]"

REQUIRED_REVIEW_COLUMNS = {"hit_id", "human_label", "span_status", "raw_text", "span_segments"}
LABEL_KEYS = ("positive", "negative")
ROLE_KEYS = ("pos_conti", "pos_disconti", "neg_target_absent")
SPLIT_KEYS = ("train", "dev", "test")

LABEL_ALIASES = {
    "tp": "tp",
    "t": "tp",
    "true_positive": "tp",
    "true positive": "tp",
    "positive": "tp",
    "pos": "tp",
    "fp": "fp",
    "f": "fp",
    "false_positive": "fp",
    "false positive": "fp",
    "negative": "fp",
    "neg": "fp",
    "unclear": "unclear",
    "unsure": "unclear",
    "unknown": "unclear",
    "?": "unclear",
}

SPAN_STATUS_ALIASES = {
    "ok": "ok",
    "o": "ok",
    "span_wrong": "span_wrong",
    "span wrong": "span_wrong",
    "wrong": "span_wrong",
    "not_applicable": "not_applicable",
    "not applicable": "not_applicable",
    "na": "not_applicable",
    "n/a": "not_applicable",
    "none": "not_applicable",
}

XLSX_COLUMNS = [
    "e_id",
    "example_id",
    "context_left",
    "target_sentence",
    "context_right",
    "instance_id",
    "split",
    "span_segments",
    "pattern_type",
    "example_role",
    "source",
    "note",
]


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _normalize_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalize_label(value: Any) -> str:
    text = _normalize_key(value)
    if not text:
        return "blank"
    return LABEL_ALIASES.get(text, "invalid")


def _normalize_span_status(value: Any) -> str:
    text = _normalize_key(value)
    if not text:
        return "blank"
    return SPAN_STATUS_ALIASES.get(text, "invalid")


def _json_cell_to_value(value: Any, default: Any) -> Any:
    if value is None:
        return default
    if isinstance(value, (list, dict)):
        return value
    text = str(value).strip()
    if not text:
        return default
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return default


def _write_json(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2, allow_nan=False)
        f.write("\n")


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"{path}: missing header row")
        normalized_fieldnames = [_normalize_header(name) for name in reader.fieldnames]
        if any(not name for name in normalized_fieldnames):
            raise ValueError(f"{path}: blank header cell inside header row")
        if len(set(normalized_fieldnames)) != len(normalized_fieldnames):
            duplicates = sorted(
                name for name in set(normalized_fieldnames) if normalized_fieldnames.count(name) > 1
            )
            raise ValueError(f"{path}: duplicate header after strip: {duplicates}")
        rows: list[dict[str, str]] = []
        for row in reader:
            rows.append(
                {
                    normalized_name: str(row.get(original_name) or "").strip()
                    for original_name, normalized_name in zip(reader.fieldnames, normalized_fieldnames)
                }
            )
        return rows


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read/write .xlsx files")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError(f"{path}: empty workbook") from None

    headers = [_normalize_header(value) for value in header_row]
    while headers and not headers[-1]:
        headers.pop()
    if not headers:
        raise ValueError(f"{path}: missing header row")
    if any(not header for header in headers):
        raise ValueError(f"{path}: blank header cell inside header row")
    if len(set(headers)) != len(headers):
        duplicates = sorted(header for header in set(headers) if headers.count(header) > 1)
        raise ValueError(f"{path}: duplicate header after strip: {duplicates}")

    rows: list[dict[str, str]] = []
    for values in rows_iter:
        trimmed_values = list(values[: len(headers)])
        if not any(value is not None and str(value).strip() for value in trimmed_values):
            continue
        row: dict[str, str] = {}
        for header, value in zip(headers, trimmed_values):
            row[header] = "" if value is None else str(value).strip()
        rows.append(row)
    return rows


def _read_review_file(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported review file extension: {path}")


def _validate_review_columns(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        raise ValueError(f"{path}: no data rows")
    columns = set(rows[0])
    missing = sorted(REQUIRED_REVIEW_COLUMNS - columns)
    if missing:
        raise ValueError(f"{path}: missing required columns: {missing}")


def _load_item_metadata(bundle_path: Path, item_id: str) -> dict[str, str]:
    bundle = json.loads(bundle_path.read_text(encoding="utf-8"))
    item = (bundle.get("items_by_e_id") or {}).get(item_id)
    if not item:
        raise ValueError(f"{bundle_path}: item_id not found in items_by_e_id: {item_id}")
    canonical_form = str(item.get("canonical_form") or "").strip()
    gloss = str(item.get("gloss") or "").strip()
    group = str(item.get("group") or "").strip()
    if not canonical_form:
        raise ValueError(f"{bundle_path}: missing canonical_form for {item_id}")
    text_b = f"{canonical_form}\n{gloss}" if gloss else canonical_form
    return {
        "canonical_form": canonical_form,
        "gloss": gloss,
        "group": group,
        "text_b": text_b,
    }


def _resolve_output_paths(
    *,
    item_id: str,
    artifact_root: Path | None,
    out_xlsx: Path | None,
    out_jsonl: Path | None,
    out_summary: Path | None,
) -> tuple[Path, Path, Path]:
    if out_xlsx and out_jsonl and out_summary:
        return out_xlsx, out_jsonl, out_summary
    if artifact_root is None:
        missing = [
            name
            for name, value in [
                ("--out-xlsx", out_xlsx),
                ("--out-jsonl", out_jsonl),
                ("--out-summary", out_summary),
            ]
            if value is None
        ]
        raise ValueError(
            f"Missing output path(s): {', '.join(missing)}. "
            "Pass them explicitly or use --artifact-root to derive item-specific paths."
        )
    item_dir = artifact_root / item_id
    return (
        out_xlsx or item_dir / f"{item_id}_encoder_examples.xlsx",
        out_jsonl or item_dir / f"{item_id}_encoder_pair_examples.jsonl",
        out_summary or item_dir / f"{item_id}_encoder_examples_summary.json",
    )


def _item_mismatch(row: dict[str, str], item_id: str) -> bool:
    origin_e_id = str(row.get("origin_e_id") or "").strip()
    unit_id = str(row.get("unit_id") or "").strip()
    if origin_e_id:
        return origin_e_id != item_id and unit_id != item_id
    if unit_id:
        return unit_id != item_id
    return False


def _make_note(row: dict[str, str], *, label: str, detect_rule_ids: list[Any], review_file: Path) -> str:
    pieces = [
        f"label={label}",
        f"source_hit_id={row.get('hit_id', '')}",
        f"batch_id={row.get('batch_id', '')}",
        f"text_id={row.get('text_id', '')}",
        f"span_source={row.get('span_source', '')}",
        f"component_span_status={row.get('component_span_status', '')}",
        f"detect_rule_ids={json.dumps(detect_rule_ids, ensure_ascii=False, separators=(',', ':'))}",
        f"review_file={review_file.name}",
    ]
    memo = str(row.get("memo") or "").strip()
    reviewer = str(row.get("reviewer") or "").strip()
    if memo:
        pieces.append(f"memo={memo}")
    if reviewer:
        pieces.append(f"reviewer={reviewer}")
    return "; ".join(pieces)


def _convert_review_row(
    *,
    row: dict[str, str],
    row_index: int,
    review_file: Path,
    item_id: str,
    item_meta: dict[str, str],
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    hit_id = str(row.get("hit_id") or "").strip()
    if not hit_id:
        raise ValueError(f"{review_file}:{row_index} blank hit_id")
    if _item_mismatch(row, item_id):
        raise ValueError(
            f"{review_file}:{row_index} item mismatch for --item-id={item_id}: "
            f"origin_e_id={row.get('origin_e_id', '')!r}, unit_id={row.get('unit_id', '')!r}"
        )

    normalized_label = _normalize_label(row.get("human_label"))
    normalized_span_status = _normalize_span_status(row.get("span_status"))
    raw_text = str(row.get("raw_text") or "").strip()

    base_skip = {
        "input_file": str(review_file),
        "row_index": row_index,
        "hit_id": hit_id,
        "human_label": row.get("human_label", ""),
        "normalized_label": normalized_label,
        "span_status": row.get("span_status", ""),
        "normalized_span_status": normalized_span_status,
    }
    if normalized_label in {"blank", "invalid", "unclear"}:
        reason = f"human_label_{normalized_label}"
        return None, {**base_skip, "reason": reason}
    if normalized_span_status == "invalid":
        return None, {**base_skip, "reason": "span_status_invalid"}
    if normalized_label == "tp" and normalized_span_status != "ok":
        return None, {**base_skip, "reason": "tp_span_status_not_ok"}
    if normalized_label == "fp" and normalized_span_status not in {"ok", "not_applicable", "blank"}:
        return None, {**base_skip, "reason": "fp_span_status_not_allowed"}
    if not raw_text:
        return None, {**base_skip, "reason": "raw_text_blank"}

    span_value = row.get("span_segments")
    span_value_source = "span_segments"
    corrected_span_segments = str(row.get("corrected_span_segments") or "").strip()
    if normalized_span_status == "span_wrong" and corrected_span_segments:
        span_value = corrected_span_segments
        span_value_source = "corrected_span_segments"

    try:
        span_segments = validate_span_segments(raw_text, parse_span_segments(span_value))
    except Exception as exc:
        return None, {**base_skip, "reason": f"{span_value_source}_invalid", "error": str(exc)}

    span_key = make_span_key(span_segments)
    span_text = make_span_text(raw_text, span_segments, gap_marker=DEFAULT_GAP_MARKER)
    pattern_type = "conti" if len(span_segments) == 1 else "disconti"
    if normalized_label == "tp":
        label = 1
        label_name = "positive"
        example_role = "pos_conti" if pattern_type == "conti" else "pos_disconti"
        negative_type = None
        id_prefix = "pos"
    else:
        label = 0
        label_name = "negative"
        example_role = "neg_target_absent"
        negative_type = "target_absent"
        id_prefix = "neg"

    detect_rule_ids = _json_cell_to_value(row.get("detect_rule_ids"), [])
    if not isinstance(detect_rule_ids, list):
        detect_rule_ids = []

    text_a = inject_span_markers(raw_text, span_segments)
    text_b = item_meta["text_b"]
    source_hit_id = hit_id
    corpus_domain = str(row.get("corpus_domain") or "").strip()
    source = str(row.get("source") or "").strip()

    record = {
        "schema_version": SCHEMA_VERSION,
        "input_construction_version": INPUT_CONSTRUCTION_VERSION,
        "span_marker_style": SPAN_MARKER_STYLE,
        "item_id": item_id,
        "example_id": "",
        "id_prefix": id_prefix,
        "label": label,
        "label_name": label_name,
        "example_role": example_role,
        "negative_type": negative_type,
        "split": "",
        "text_a": text_a,
        "text_b": text_b,
        "raw_text": raw_text,
        "target_sentence": raw_text,
        "span_segments": span_segments,
        "span_key": span_key,
        "span_text": span_text,
        "span_value_source": span_value_source,
        "canonical_form": item_meta["canonical_form"],
        "gloss": item_meta["gloss"],
        "group": item_meta["group"],
        "pattern_type": pattern_type,
        "source_hit_id": source_hit_id,
        "hit_id": hit_id,
        "batch_id": str(row.get("batch_id") or ""),
        "text_id": str(row.get("text_id") or ""),
        "corpus_domain": corpus_domain,
        "source": source,
        "source_file": str(row.get("source_file") or ""),
        "source_row_index": str(row.get("source_row_index") or ""),
        "source_line_no": str(row.get("source_line_no") or ""),
        "regex_match_text": str(row.get("regex_match_text") or ""),
        "span_source": str(row.get("span_source") or ""),
        "component_span_status": str(row.get("component_span_status") or ""),
        "detect_rule_ids": detect_rule_ids,
        "human_label": normalized_label,
        "span_status": normalized_span_status,
        "memo": str(row.get("memo") or ""),
        "reviewer": str(row.get("reviewer") or ""),
        "review_file": str(review_file),
        "note": _make_note(row, label=normalized_label, detect_rule_ids=detect_rule_ids, review_file=review_file),
    }
    return record, None


def _stable_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _assign_splits(
    records: list[dict[str, Any]],
    *,
    item_id: str,
    seed: int,
    train_ratio: float = 0.8,
    dev_ratio: float = 0.1,
    test_ratio: float = 0.1,
) -> list[str]:
    warnings: list[str] = []
    by_role: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_role[str(record["example_role"])].append(record)

    for role, role_records in sorted(by_role.items()):
        role_records.sort(
            key=lambda record: _stable_hash(
                "|".join(
                    [
                        str(seed),
                        item_id,
                        role,
                        str(record.get("hit_id") or ""),
                        str(record.get("raw_text") or ""),
                        str(record.get("span_key") or ""),
                    ]
                )
            )
        )
        n = len(role_records)
        if n >= 10:
            n_dev = max(1, round(n * dev_ratio))
            n_test = max(1, round(n * test_ratio))
            if n_dev + n_test >= n:
                n_dev = 1
                n_test = 1
            n_train = n - n_dev - n_test
        else:
            n_train, n_dev, n_test = n, 0, 0
            warnings.append(f"role {role} has fewer than 10 examples; assigned all {n} to train")

        for idx, record in enumerate(role_records):
            if idx < n_train:
                record["split"] = "train"
            elif idx < n_train + n_dev:
                record["split"] = "dev"
            else:
                record["split"] = "test"
    return warnings


def _deduplicate_records(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    seen: dict[tuple[str, int, str, str], dict[str, Any]] = {}
    deduped: list[dict[str, Any]] = []
    dropped: list[dict[str, str]] = []
    for record in records:
        key = (
            str(record["item_id"]),
            int(record["label"]),
            str(record["raw_text"]),
            str(record["span_key"]),
        )
        if key in seen:
            dropped.append(
                {
                    "kept_hit_id": str(seen[key].get("hit_id") or ""),
                    "dropped_hit_id": str(record.get("hit_id") or ""),
                    "raw_text": str(record.get("raw_text") or ""),
                    "span_key": str(record.get("span_key") or ""),
                }
            )
            continue
        seen[key] = record
        deduped.append(record)
    return deduped, dropped


def _assign_example_ids(records: list[dict[str, Any]], item_id: str) -> None:
    records.sort(key=lambda record: (0 if record["label"] == 1 else 1, str(record.get("hit_id") or "")))
    counters = {"pos": 0, "neg": 0}
    for record in records:
        prefix = str(record["id_prefix"])
        counters[prefix] += 1
        record["example_id"] = f"{item_id}-{prefix}-{counters[prefix]:04d}"


def _counter_by(records: list[dict[str, Any]], key: str) -> dict[str, int]:
    return dict(sorted(Counter(str(record.get(key) or "unknown") for record in records).items()))


def _split_counts_by_role(records: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = {}
    for role in ROLE_KEYS:
        split_counter = Counter(
            str(record.get("split") or "unknown")
            for record in records
            if str(record.get("example_role") or "") == role
        )
        result[role] = {split: int(split_counter.get(split, 0)) for split in SPLIT_KEYS}
    return result


def _write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            out_record = {key: value for key, value in record.items() if key != "id_prefix"}
            f.write(json.dumps(out_record, ensure_ascii=False, allow_nan=False) + "\n")


def _xlsx_row(record: dict[str, Any]) -> dict[str, str]:
    source = str(record.get("corpus_domain") or record.get("source") or "")
    return {
        "e_id": str(record["item_id"]),
        "example_id": str(record["example_id"]),
        "context_left": "",
        "target_sentence": str(record["target_sentence"]),
        "context_right": "",
        "instance_id": "1",
        "split": str(record["split"]),
        "span_segments": format_span_segments(record["span_segments"]),
        "pattern_type": str(record["pattern_type"]),
        "example_role": str(record["example_role"]),
        "source": source,
        "note": str(record.get("note") or ""),
    }


def _write_xlsx(path: Path, records: list[dict[str, Any]]) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to write .xlsx files")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "examples"
    sheet.append(XLSX_COLUMNS)
    for record in records:
        row = _xlsx_row(record)
        sheet.append([row[column] for column in XLSX_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 12,
        "B": 18,
        "D": 70,
        "F": 18,
        "G": 10,
        "H": 22,
        "I": 12,
        "J": 22,
        "K": 18,
        "L": 100,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)


def export_encoder_examples(
    *,
    item_id: str,
    bundle_path: Path,
    input_paths: list[Path],
    out_xlsx: Path,
    out_jsonl: Path,
    out_summary: Path,
    min_pos: int,
    min_neg: int,
    seed: int,
) -> dict[str, Any]:
    if not item_id.strip():
        raise ValueError("--item-id must not be blank")
    if not input_paths:
        raise ValueError("At least one --input is required")

    expected_item_id = item_id.strip()
    item_meta = _load_item_metadata(bundle_path, expected_item_id)
    raw_records: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    n_rows_read = 0
    seen_hit_ids: dict[str, str] = {}

    for input_path in input_paths:
        rows = _read_review_file(input_path)
        _validate_review_columns(input_path, rows)
        for row_index, row in enumerate(rows, start=2):
            hit_id = str(row.get("hit_id") or "").strip()
            if not hit_id:
                raise ValueError(f"{input_path}:{row_index} blank hit_id")
            if hit_id in seen_hit_ids:
                raise ValueError(
                    f"Duplicate hit_id found: {hit_id} first_file={seen_hit_ids[hit_id]} duplicate_file={input_path}"
                )
            seen_hit_ids[hit_id] = str(input_path)
            n_rows_read += 1
            record, skipped = _convert_review_row(
                row=row,
                row_index=row_index,
                review_file=input_path,
                item_id=expected_item_id,
                item_meta=item_meta,
            )
            if record is not None:
                raw_records.append(record)
            elif skipped is not None:
                skipped_rows.append(skipped)

    deduped_records, deduped_examples = _deduplicate_records(raw_records)
    _assign_example_ids(deduped_records, expected_item_id)
    split_warnings = _assign_splits(deduped_records, item_id=expected_item_id, seed=seed)

    label_counts = Counter("positive" if record["label"] == 1 else "negative" for record in deduped_records)
    role_counts = Counter(str(record["example_role"]) for record in deduped_records)
    split_counts = Counter(str(record["split"]) for record in deduped_records)
    warnings = list(split_warnings)
    if label_counts.get("positive", 0) < min_pos:
        warnings.append(f"positive count below min_pos={min_pos}")
    if label_counts.get("negative", 0) < min_neg:
        warnings.append(f"negative count below min_neg={min_neg}")

    _write_jsonl(out_jsonl, deduped_records)
    _write_xlsx(out_xlsx, deduped_records)

    summary = {
        "schema_version": SUMMARY_SCHEMA_VERSION,
        "item_id": expected_item_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "input_files": [str(path) for path in input_paths],
        "bundle": str(bundle_path),
        "out_xlsx": str(out_xlsx),
        "out_jsonl": str(out_jsonl),
        "n_rows_read": n_rows_read,
        "n_rows_exported": len(deduped_records),
        "n_rows_skipped": len(skipped_rows),
        "skipped_counts_by_reason": dict(sorted(Counter(row["reason"] for row in skipped_rows).items())),
        "skipped_rows": skipped_rows[:100],
        "n_skipped_rows_listed": min(len(skipped_rows), 100),
        "deduped_count": len(deduped_examples),
        "deduped_examples": deduped_examples[:100],
        "n_deduped_examples_listed": min(len(deduped_examples), 100),
        "label_counts": {key: int(label_counts.get(key, 0)) for key in LABEL_KEYS},
        "role_counts": {key: int(role_counts.get(key, 0)) for key in ROLE_KEYS},
        "split_counts": {key: int(split_counts.get(key, 0)) for key in SPLIT_KEYS},
        "split_counts_by_role": _split_counts_by_role(deduped_records),
        "counts_by_domain": _counter_by(deduped_records, "corpus_domain"),
        "target_reached": {
            "positive_100": int(label_counts.get("positive", 0)) >= min_pos,
            "negative_100": int(label_counts.get("negative", 0)) >= min_neg,
        },
        "ready_for_training": int(label_counts.get("positive", 0)) >= min_pos
        and int(label_counts.get("negative", 0)) >= min_neg,
        "span_format": "json_list",
        "input_construction_version": INPUT_CONSTRUCTION_VERSION,
        "span_marker_style": SPAN_MARKER_STYLE,
        "split_policy": {
            "seed": seed,
            "roles": list(ROLE_KEYS),
            "train_dev_test": [0.8, 0.1, 0.1],
            "small_role_policy": "if role_count < 10, assign all to train and warn",
            "hash_payload": "seed|item_id|example_role|hit_id|raw_text|span_key",
        },
        "warnings": warnings,
    }
    _write_json(out_summary, summary)
    return summary


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--item-id", required=True, help="Grammar item id, e.g. df003")
    parser.add_argument("--bundle", required=True, type=Path, help="Detector bundle JSON path")
    parser.add_argument(
        "--input",
        action="append",
        required=True,
        dest="inputs",
        help="Human-labeled review .xlsx or .csv file. Can be passed multiple times.",
    )
    parser.add_argument("--out-xlsx", type=Path)
    parser.add_argument("--out-jsonl", type=Path)
    parser.add_argument("--out-summary", type=Path)
    parser.add_argument(
        "--artifact-root",
        type=Path,
        help="Base artifact folder, e.g. /.../HanTalk_arti/example_making. Missing output paths are written under {artifact_root}/{item_id}/.",
    )
    parser.add_argument("--min-pos", type=int, default=100)
    parser.add_argument("--min-neg", type=int, default=100)
    parser.add_argument("--seed", type=int, default=20260502)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)
    try:
        out_xlsx, out_jsonl, out_summary = _resolve_output_paths(
            item_id=args.item_id.strip(),
            artifact_root=args.artifact_root,
            out_xlsx=args.out_xlsx,
            out_jsonl=args.out_jsonl,
            out_summary=args.out_summary,
        )
        summary = export_encoder_examples(
            item_id=args.item_id,
            bundle_path=args.bundle,
            input_paths=[Path(path) for path in args.inputs],
            out_xlsx=out_xlsx,
            out_jsonl=out_jsonl,
            out_summary=out_summary,
            min_pos=args.min_pos,
            min_neg=args.min_neg,
            seed=args.seed,
        )
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1

    print(
        json.dumps(
            {
                "item_id": summary["item_id"],
                "n_rows_read": summary["n_rows_read"],
                "n_rows_exported": summary["n_rows_exported"],
                "label_counts": summary["label_counts"],
                "role_counts": summary["role_counts"],
                "split_counts": summary["split_counts"],
                "target_reached": summary["target_reached"],
                "ready_for_training": summary["ready_for_training"],
                "out_xlsx": summary["out_xlsx"],
                "out_jsonl": summary["out_jsonl"],
                "out_summary": str(out_summary),
            },
            ensure_ascii=False,
            indent=2,
            allow_nan=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
